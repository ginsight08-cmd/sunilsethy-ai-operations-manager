"""
Procurement / Vendor Price Comparison engine — Manufacturing industry module
for AI Operations Manager.

Supports TWO workbook formats, auto-detected:

1. PRICE COMPARISON — "Price Comparative Sheet" style workbooks (one sheet
   per Purchase Requisition, N vendor quote blocks per sheet, plus a
   Previous Order Details block). Produces vendor-comparison analysis:
   recommended vendor per item, savings vs alternatives, price trend vs
   the last order, and risk flags.

2. COMMERCIAL COMPARISON — vendor/commercial-terms comparison workbooks
   (Make, Freight, Delivery, Payment, Warranty, Previous Order Date, etc.)
   with no price fields. Produces sourcing/risk/data-quality analysis
   instead of spend/savings metrics.

This is intentionally a SEPARATE module from engine.py (the BPO/ops-KPI
engine) — Manufacturing procurement analysis has nothing in common with
Productivity/Quality/SLA/AHT, so it doesn't touch or depend on engine.py.
"""

import hashlib
import json
import re
from datetime import datetime

import openpyxl
import pandas as pd


# ============================================================
# SHARED HELPERS
# ============================================================

def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ============================================================
# FORMAT DETECTION
# ============================================================

def detect_workbook_format(filepath):
    """
    Detect the Manufacturing workbook structure.

    Returns one of: "price_comparison", "commercial_comparison", "unsupported"
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)

    price_score = 0
    commercial_score = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            values = [_clean(value).lower() for value in row if _clean(value)]
            joined = " | ".join(values)

            if "quoted price" in joined:
                price_score += 2
            if "total price" in joined:
                price_score += 2
            if "item code" in joined:
                price_score += 1
            if "item name" in joined:
                price_score += 1
            if "previous order details" in joined:
                commercial_score += 1
            if "freight" in joined:
                commercial_score += 2
            if "application" in joined:
                commercial_score += 1
            if "prepared by" in joined:
                commercial_score += 1

    if price_score >= 4:
        return "price_comparison"
    if commercial_score >= 3:
        return "commercial_comparison"
    return "unsupported"


# ============================================================
# FORMAT 1: PRICE COMPARISON PARSER
# ============================================================

def _find_header_row(ws, max_scan=20):
    for r in range(1, max_scan + 1):
        val = _clean(ws.cell(row=r, column=1).value)
        if val == "S.N.":
            return r
    return None


def _find_vendor_row(ws, header_row):
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            if "previous order details" in _clean(ws.cell(row=r, column=c).value).lower():
                return r
    return header_row - 3 if header_row and header_row > 3 else None


def _parse_sheet(ws, sheet_name):
    header_row = _find_header_row(ws)
    if header_row is None:
        return None  # not a Price Comparative Sheet — skip silently

    vendor_row = _find_vendor_row(ws, header_row)
    max_col = ws.max_column

    headers = {c: _clean(ws.cell(row=header_row, column=c).value) for c in range(1, max_col + 1)}

    col_sn, col_prno, col_itemcode, col_itemname, col_qty, col_um = 1, 2, 3, 4, 5, 6

    vendor_blocks = []
    c = 7
    while c <= max_col:
        label = headers.get(c, "")
        if label == "Doc. Date":
            break
        if label.startswith("Make"):
            vendor_name = _clean(ws.cell(row=vendor_row, column=c).value) if vendor_row else f"Vendor@{c}"
            vendor_blocks.append({
                "name": vendor_name or f"Vendor@{c}",
                "make_col": c,
                "quoted_col": c + 1,
                "disc_pct_col": c + 2,
                "disc_price_col": c + 3,
                "total_col": c + 4,
            })
            c += 5
        else:
            c += 1

    doc_date_col = vendor_col = dptpl_col = None
    for c2 in range(c, max_col + 1):
        label = headers.get(c2, "")
        if label == "Doc. Date":
            doc_date_col = c2
        elif label == "Vendor":
            vendor_col = c2
        elif label == "DPTPL":
            dptpl_col = c2

    plant_indenter = ""
    working_date = ""
    for r in range(1, header_row):
        text = _clean(ws.cell(row=r, column=1).value)
        if text.lower().startswith("plant"):
            plant_indenter = text
        if "working date" in text.lower():
            m = re.search(r"working date\s*:?\s*([0-9]{1,2}[-.][0-9]{1,2}[-.][0-9]{2,4})", text, re.IGNORECASE)
            if m:
                working_date = m.group(1)

    items = []
    r = header_row + 1
    while r <= ws.max_row:
        sn_val = ws.cell(row=r, column=col_sn).value
        if not _is_number(sn_val):
            break

        item_code = _clean(ws.cell(row=r, column=col_itemcode).value)
        item_name = _clean(ws.cell(row=r, column=col_itemname).value)
        qty = ws.cell(row=r, column=col_qty).value
        um = _clean(ws.cell(row=r, column=col_um).value)
        pr_no = _clean(ws.cell(row=r, column=col_prno).value)

        quotes = []
        for vb in vendor_blocks:
            total = ws.cell(row=r, column=vb["total_col"]).value
            unit = ws.cell(row=r, column=vb["disc_price_col"]).value
            if _is_number(total) and total > 0:
                quotes.append({
                    "vendor": vb["name"],
                    "unit_price": unit if _is_number(unit) else None,
                    "total_price": total,
                })

        prev_price = ws.cell(row=r, column=dptpl_col).value if dptpl_col else None
        prev_vendor = _clean(ws.cell(row=r, column=vendor_col).value) if vendor_col else ""
        prev_date = _clean(ws.cell(row=r, column=doc_date_col).value) if doc_date_col else ""

        items.append({
            "pr_no": pr_no,
            "item_code": item_code,
            "item_name": item_name,
            "qty": qty,
            "um": um,
            "quotes": quotes,
            "prev_unit_price": prev_price if _is_number(prev_price) else None,
            "prev_vendor": prev_vendor,
            "prev_date": prev_date,
        })
        r += 1

    return {
        "sheet_name": sheet_name,
        "plant_indenter": plant_indenter,
        "working_date": working_date,
        "vendor_count": len(vendor_blocks),
        "items": items,
    }


def parse_workbook(filepath):
    """Parse every Price-Comparative-Sheet-style sheet in the workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prs = []
    for sn in wb.sheetnames:
        parsed = _parse_sheet(wb[sn], sn)
        if parsed and parsed["items"]:
            prs.append(parsed)
    return prs


def analyze_procurement(prs, price_increase_risk_pct=10.0):
    """
    Turn parsed PR sheets into item-level and PR-level analysis.
    Returns dict with 'item_rows', 'pr_rows', 'overall'.
    """
    item_rows = []

    for pr in prs:
        for it in pr["items"]:
            quotes = sorted(it["quotes"], key=lambda q: q["total_price"])
            n_quotes = len(quotes)

            if n_quotes == 0:
                item_rows.append({
                    "PR Sheet": pr["sheet_name"],
                    "PR No": it["pr_no"],
                    "Item Code": it["item_code"],
                    "Item Name": it["item_name"],
                    "Qty": it["qty"],
                    "UM": it["um"],
                    "Vendors Quoted": 0,
                    "Recommended Vendor": None,
                    "Recommended Total Price": None,
                    "2nd Best Vendor": None,
                    "2nd Best Total Price": None,
                    "Highest Quoted Vendor": None,
                    "Highest Total Price": None,
                    "Savings vs Highest": None,
                    "Savings vs Highest %": None,
                    "Previous Unit Price": it["prev_unit_price"],
                    "Previous Vendor": it["prev_vendor"],
                    "% Change vs Previous": None,
                    "Risk Flags": "No vendor quote available",
                })
                continue

            best = quotes[0]
            worst = quotes[-1]
            second = quotes[1] if n_quotes > 1 else None

            savings_vs_highest = worst["total_price"] - best["total_price"]
            savings_vs_highest_pct = (
                (savings_vs_highest / worst["total_price"] * 100.0) if worst["total_price"] else None
            )

            pct_change_vs_prev = None
            if it["prev_unit_price"] and best["unit_price"] is not None and it["prev_unit_price"] > 0:
                pct_change_vs_prev = (
                    (best["unit_price"] - it["prev_unit_price"]) / it["prev_unit_price"] * 100.0
                )

            risk_flags = []
            if n_quotes == 1:
                risk_flags.append("Single vendor quoted")
            if pct_change_vs_prev is not None and pct_change_vs_prev >= price_increase_risk_pct:
                risk_flags.append(f"Price up {pct_change_vs_prev:.1f}% vs last order")

            item_rows.append({
                "PR Sheet": pr["sheet_name"],
                "PR No": it["pr_no"],
                "Item Code": it["item_code"],
                "Item Name": it["item_name"],
                "Qty": it["qty"],
                "UM": it["um"],
                "Vendors Quoted": n_quotes,
                "Recommended Vendor": best["vendor"],
                "Recommended Total Price": round(best["total_price"], 2),
                "2nd Best Vendor": second["vendor"] if second else None,
                "2nd Best Total Price": round(second["total_price"], 2) if second else None,
                "Highest Quoted Vendor": worst["vendor"],
                "Highest Total Price": round(worst["total_price"], 2),
                "Savings vs Highest": round(savings_vs_highest, 2),
                "Savings vs Highest %": round(savings_vs_highest_pct, 1) if savings_vs_highest_pct is not None else None,
                "Previous Unit Price": it["prev_unit_price"],
                "Previous Vendor": it["prev_vendor"],
                "% Change vs Previous": round(pct_change_vs_prev, 1) if pct_change_vs_prev is not None else None,
                "Risk Flags": "; ".join(risk_flags) if risk_flags else "",
            })

    pr_rows = []
    for pr in prs:
        rows = [r for r in item_rows if r["PR Sheet"] == pr["sheet_name"]]
        recommended_spend = sum(r["Recommended Total Price"] or 0 for r in rows)
        highest_spend = sum(r["Highest Total Price"] or 0 for r in rows)
        savings = highest_spend - recommended_spend
        single_vendor_items = sum(1 for r in rows if "Single vendor" in (r["Risk Flags"] or ""))
        price_increase_items = sum(1 for r in rows if "Price up" in (r["Risk Flags"] or ""))
        no_quote_items = sum(1 for r in rows if r["Vendors Quoted"] == 0)

        pr_rows.append({
            "PR Sheet": pr["sheet_name"],
            "Plant / Indenter": pr["plant_indenter"],
            "Working Date": pr["working_date"],
            "Vendors Compared": pr["vendor_count"],
            "Items": len(rows),
            "Recommended Spend": round(recommended_spend, 2),
            "Highest-Quote Spend": round(highest_spend, 2),
            "Potential Savings": round(savings, 2),
            "Savings %": round(savings / highest_spend * 100.0, 1) if highest_spend else None,
            "Single-Vendor Items (Risk)": single_vendor_items,
            "Price-Increase Items (Risk)": price_increase_items,
            "No-Quote Items (Risk)": no_quote_items,
        })

    total_recommended = sum(r["Recommended Spend"] for r in pr_rows)
    total_highest = sum(r["Highest-Quote Spend"] for r in pr_rows)
    overall = {
        "total_prs": len(pr_rows),
        "total_items": sum(r["Items"] for r in pr_rows),
        "total_recommended_spend": round(total_recommended, 2),
        "total_highest_quote_spend": round(total_highest, 2),
        "total_potential_savings": round(total_highest - total_recommended, 2),
        "overall_savings_pct": round((total_highest - total_recommended) / total_highest * 100.0, 1) if total_highest else None,
        "total_single_vendor_items": sum(r["Single-Vendor Items (Risk)"] for r in pr_rows),
        "total_price_increase_items": sum(r["Price-Increase Items (Risk)"] for r in pr_rows),
        "total_no_quote_items": sum(r["No-Quote Items (Risk)"] for r in pr_rows),
    }

    return {"item_rows": item_rows, "pr_rows": pr_rows, "overall": overall}


# ============================================================
# PRICE-COMPARISON INSIGHT LAYER
# ============================================================

def compute_risk_level(overall):
    breaches = sum([
        overall["total_single_vendor_items"] > 0,
        overall["total_price_increase_items"] > 0,
        overall["total_no_quote_items"] > 0,
    ])
    level = {
        0: "🟢 LOW RISK",
        1: "🟡 MEDIUM RISK",
        2: "🟠 HIGH RISK",
    }.get(breaches, "🔴 CRITICAL RISK")
    return level, breaches


def build_summary_points(overall):
    return [
        f"{'🟢' if overall['total_potential_savings'] > 0 else '⚪'} "
        f"Potential savings: ₹{overall['total_potential_savings']:,.0f} "
        f"({overall['overall_savings_pct']}%) identified across "
        f"{overall['total_prs']} purchase requisition(s).",

        f"{'🔴' if overall['total_single_vendor_items'] > 0 else '🟢'} "
        f"Single-vendor items: {overall['total_single_vendor_items']} — "
        "no competitive comparison available for these purchases.",

        f"{'🟠' if overall['total_price_increase_items'] > 0 else '🟢'} "
        f"Price-increase items: {overall['total_price_increase_items']} — "
        "priced 10%+ higher than the last recorded order.",

        f"{'🔴' if overall['total_no_quote_items'] > 0 else '🟢'} "
        f"No-quote items: {overall['total_no_quote_items']} — no valid "
        "vendor quote was received.",
    ]


def build_recommendation(breaches):
    if breaches >= 3:
        return (
            "Immediate procurement review is recommended. Multiple risk "
            "categories are present across this batch — prioritize sourcing "
            "additional vendor quotes and validating the flagged price "
            "increases before approving these purchase requisitions."
        )
    if breaches >= 1:
        return (
            "Procurement should review the flagged items, confirm pricing "
            "with the recommended vendors, and source a second quote for "
            "any single-vendor items where feasible."
        )
    return (
        "All items have competitive quotes with no unusual price movement. "
        "Proceed with the recommended vendors listed in the Item Detail report."
    )


def build_insights(result):
    overall = result["overall"]
    risk_level, breaches = compute_risk_level(overall)
    summary_points = build_summary_points(overall)
    recommendation = build_recommendation(breaches)
    return {
        "risk_level": risk_level,
        "breaches": breaches,
        "summary_points": summary_points,
        "recommendation": recommendation,
    }


def make_ai_prompt(result, insights=None):
    overall = result["overall"]
    insights = insights or build_insights(result)

    item_lines = []
    for row in result["item_rows"][:200]:
        item_lines.append(
            f"- {row['Item Name']} (PR {row['PR Sheet']}, Qty {row['Qty']} {row['UM']}): "
            f"recommended {row['Recommended Vendor']} at ₹{row['Recommended Total Price']}, "
            f"highest quote ₹{row['Highest Total Price']}"
            + (f", 2nd best {row['2nd Best Vendor']} at ₹{row['2nd Best Total Price']}" if row['2nd Best Vendor'] else "")
            + (f", prev. unit price ₹{row['Previous Unit Price']}" if row['Previous Unit Price'] is not None else "")
            + (f", risk: {row['Risk Flags']}" if row['Risk Flags'] else "")
        )

    pr_lines = [
        f"- {row['PR Sheet']} ({row['Plant / Indenter'] or 'plant n/a'}, {row['Working Date'] or 'date n/a'}): "
        f"{row['Items']} items, {row['Vendors Compared']} vendors compared, "
        f"₹{row['Recommended Spend']:,.0f} recommended spend, "
        f"₹{row['Potential Savings']:,.0f} potential savings"
        for row in result["pr_rows"]
    ]

    return f"""
PROCUREMENT PRICE COMPARISON — ANALYSIS CONTEXT

Overall Risk: {insights['risk_level']}

Overview:
- Purchase Requisitions: {overall['total_prs']}
- Total Items: {overall['total_items']}
- Recommended Spend: ₹{overall['total_recommended_spend']:,.0f}
- Highest-Quote Spend: ₹{overall['total_highest_quote_spend']:,.0f}
- Potential Savings: ₹{overall['total_potential_savings']:,.0f} ({overall['overall_savings_pct']}%)
- Single-Vendor Items: {overall['total_single_vendor_items']}
- Price-Increase Items: {overall['total_price_increase_items']}
- No-Quote Items: {overall['total_no_quote_items']}

Executive Summary:
{chr(10).join('- ' + p for p in insights['summary_points'])}

Recommendation:
{insights['recommendation']}

Purchase Requisition Breakdown:
{chr(10).join(pr_lines)}

Item-Level Detail:
{chr(10).join(item_lines)}
""".strip()


# ============================================================
# FORMAT 2: COMMERCIAL COMPARISON PARSER
# ============================================================

COMMERCIAL_TERM_NAMES = {
    1: "Freight Terms",
    2: "Taxes",
    3: "Other Charges",
    4: "Delivery Period",
    5: "Payment Terms",
    6: "Payment Basis",
    7: "Warranty",
    8: "Term 8",
    9: "Term 9",
    10: "Communication Mode",
}


def _normalize_name(value):
    value = _clean(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    text = _clean(value)
    for fmt in ("%d-%m-%Y", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d.%m.%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _extract_working_date(value):
    text = _clean(value)
    match = re.search(
        r"working\s*date\s*:?\s*(\d{1,2}[-./]\d{1,2}[-./]\d{2,4})",
        text, flags=re.IGNORECASE,
    )
    if not match:
        return None
    return _parse_date(match.group(1))


def _commercial_sheet_fingerprint(parsed):
    payload = {
        "working_date": parsed["working_date"].isoformat() if parsed["working_date"] else "",
        "vendors": [_normalize_name(v["vendor_name"]) for v in parsed["vendors"]],
        "items": [
            {
                "sn": item["sn"],
                "makes": item["makes"],
                "previous_order_date": str(item.get("previous_order_date") or ""),
            }
            for item in parsed["items"]
        ],
        "prepared_by": _normalize_name(parsed["prepared_by"]),
    }
    encoded = json.dumps(payload, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _find_row_by_first_cell(ws, expected):
    expected = expected.lower()
    for row_number in range(1, ws.max_row + 1):
        value = _clean(ws.cell(row=row_number, column=1).value).lower()
        if value == expected:
            return row_number
    return None


def _parse_commercial_sheet(ws, sheet_name):
    header_row = _find_row_by_first_cell(ws, "s.n.")
    if not header_row:
        return None

    previous_order_col = None
    for column in range(1, ws.max_column + 1):
        for row in range(1, header_row + 1):
            value = _clean(ws.cell(row=row, column=column).value).lower()
            if "previous order details" in value:
                previous_order_col = column
                break
        if previous_order_col:
            break
    if not previous_order_col:
        previous_order_col = ws.max_column

    vendor_row = None
    for row in range(1, header_row):
        non_empty = [_clean(ws.cell(row=row, column=c).value) for c in range(2, previous_order_col)]
        if sum(bool(value) for value in non_empty) >= 1:
            vendor_row = row
    if not vendor_row:
        vendor_row = 1

    working_date = None
    for row in range(1, header_row):
        for column in range(1, ws.max_column + 1):
            detected = _extract_working_date(ws.cell(row=row, column=column).value)
            if detected:
                working_date = detected
                break
        if working_date:
            break

    vendors = []
    for column in range(2, previous_order_col):
        vendor_name = _clean(ws.cell(row=vendor_row, column=column).value)
        if not vendor_name:
            continue
        if vendor_name.lower() in {"contact no.& name", "make"}:
            continue
        vendors.append({
            "vendor_name": vendor_name,
            "normalized_name": _normalize_name(vendor_name),
            "column": column,
            "contact": "",
            "commercial_terms": {},
        })

    freight_row = _find_row_by_first_cell(ws, "freight")

    items = []
    row = header_row + 1
    while row <= ws.max_row:
        sn = ws.cell(row=row, column=1).value
        if not isinstance(sn, (int, float)):
            break

        makes = {}
        for vendor in vendors:
            makes[vendor["vendor_name"]] = _clean(ws.cell(row=row, column=vendor["column"]).value)

        previous_value = ws.cell(row=row, column=previous_order_col).value if previous_order_col else None

        items.append({
            "sn": int(sn),
            "makes": makes,
            "previous_order_date": _parse_date(previous_value),
            "previous_order_raw": _clean(previous_value),
        })
        row += 1

    if freight_row:
        row = freight_row + 1
        while row <= ws.max_row:
            term_no = ws.cell(row=row, column=1).value
            if not isinstance(term_no, (int, float)):
                break
            term_no = int(term_no)
            term_name = COMMERCIAL_TERM_NAMES.get(term_no, f"Term {term_no}")
            for vendor in vendors:
                vendor["commercial_terms"][term_name] = _clean(ws.cell(row=row, column=vendor["column"]).value)
            row += 1

    prepared_by = ""
    for row in range(1, ws.max_row + 1):
        value = _clean(ws.cell(row=row, column=1).value)
        if value.lower().startswith("prepared by"):
            prepared_by = value.split(":", 1)[-1].strip()
            break

    parsed = {
        "sheet_name": sheet_name,
        "working_date": working_date,
        "vendors": vendors,
        "items": items,
        "prepared_by": prepared_by,
    }
    parsed["fingerprint"] = _commercial_sheet_fingerprint(parsed)
    return parsed


def parse_commercial_workbook(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    cases = []
    for sheet_name in wb.sheetnames:
        parsed = _parse_commercial_sheet(wb[sheet_name], sheet_name)
        if parsed and parsed["items"]:
            cases.append(parsed)
    return cases


def analyze_commercial_procurement(cases, analysis_date=None, stale_days=365):
    analysis_date = analysis_date or datetime.now().date()

    seen_fingerprints = {}
    duplicate_sheets = []
    case_rows = []
    item_rows = []
    vendor_rows = []
    distinct_cases = []

    for case in cases:
        fingerprint = case["fingerprint"]
        if fingerprint in seen_fingerprints:
            duplicate_sheets.append({
                "Duplicate Sheet": case["sheet_name"],
                "Original Sheet": seen_fingerprints[fingerprint],
                "Reason": "Identical procurement content",
            })
            continue
        seen_fingerprints[fingerprint] = case["sheet_name"]
        distinct_cases.append(case)

    total_items = 0
    total_vendor_columns = 0
    single_source_cases = 0
    stale_references = 0
    recent_references = 0
    new_item_references = 0
    missing_previous_dates = 0
    missing_warranty_quotes = 0
    extended_lead_time_quotes = 0
    duplicate_vendor_cases = 0

    for case in distinct_cases:
        vendors = case["vendors"]
        items = case["items"]
        total_items += len(items)
        total_vendor_columns += len(vendors)

        normalized_names = [v["normalized_name"] for v in vendors if v["normalized_name"]]
        duplicated_vendors = sorted({n for n in normalized_names if normalized_names.count(n) > 1})
        if duplicated_vendors:
            duplicate_vendor_cases += 1
        if len(set(normalized_names)) <= 1:
            single_source_cases += 1

        stale_in_case = 0
        new_in_case = 0
        missing_previous_in_case = 0
        warranty_missing_in_case = 0
        extended_delivery_in_case = 0

        for item in items:
            previous_date = item["previous_order_date"]
            previous_raw = item["previous_order_raw"]
            age_days = None
            recency = "Unavailable"

            if previous_date:
                age_days = (analysis_date - previous_date).days
                if age_days > stale_days:
                    recency = "Stale"
                    stale_references += 1
                    stale_in_case += 1
                else:
                    recency = "Current"
                    recent_references += 1
            elif "new item" in previous_raw.lower():
                recency = "New Item"
                new_item_references += 1
                new_in_case += 1
            else:
                missing_previous_dates += 1
                missing_previous_in_case += 1

            item_rows.append({
                "Sheet": case["sheet_name"],
                "Working Date": case["working_date"],
                "S.N.": item["sn"],
                "Vendor Makes": "; ".join(f"{v}: {m or 'Missing'}" for v, m in item["makes"].items()),
                "Previous Order Date": previous_date,
                "Reference Age Days": age_days,
                "Reference Status": recency,
            })

        for vendor in vendors:
            terms = vendor["commercial_terms"]
            warranty = terms.get("Warranty", "")
            delivery = terms.get("Delivery Period", "")
            if not warranty:
                missing_warranty_quotes += 1
                warranty_missing_in_case += 1
            delivery_lower = delivery.lower()
            is_extended = "week" in delivery_lower and not ("7-10 day" in delivery_lower or "7–10 day" in delivery_lower)
            if is_extended:
                extended_lead_time_quotes += 1
                extended_delivery_in_case += 1

            vendor_rows.append({
                "Sheet": case["sheet_name"],
                "Vendor": vendor["vendor_name"],
                "Freight Terms": terms.get("Freight Terms", ""),
                "Taxes": terms.get("Taxes", ""),
                "Other Charges": terms.get("Other Charges", ""),
                "Delivery Period": delivery,
                "Payment Terms": terms.get("Payment Terms", ""),
                "Payment Basis": terms.get("Payment Basis", ""),
                "Warranty": warranty,
                "Communication Mode": terms.get("Communication Mode", ""),
            })

        risks = []
        if len(set(normalized_names)) <= 1:
            risks.append("Single-source procurement")
        if stale_in_case:
            risks.append(f"{stale_in_case} stale previous-order reference(s)")
        if new_in_case:
            risks.append(f"{new_in_case} new-item reference(s)")
        if warranty_missing_in_case:
            risks.append(f"{warranty_missing_in_case} quotation(s) missing warranty")
        if extended_delivery_in_case:
            risks.append(f"{extended_delivery_in_case} extended lead-time quotation(s)")
        if duplicated_vendors:
            risks.append("Duplicate supplier entry: " + ", ".join(duplicated_vendors))

        risk_score = 0
        if len(set(normalized_names)) <= 1:
            risk_score += 20
        if stale_in_case:
            risk_score += 15
        if new_in_case:
            risk_score += 20
        if warranty_missing_in_case:
            risk_score += 10
        if extended_delivery_in_case:
            risk_score += 10
        if duplicated_vendors:
            risk_score += 15
        risk_score = min(risk_score, 100)

        if risk_score >= 60:
            risk_level = "Critical"
        elif risk_score >= 40:
            risk_level = "High"
        elif risk_score >= 20:
            risk_level = "Moderate"
        else:
            risk_level = "Low"

        case_rows.append({
            "Sheet": case["sheet_name"],
            "Working Date": case["working_date"],
            "Vendor Columns": len(vendors),
            "Distinct Vendors": len(set(normalized_names)),
            "Items": len(items),
            "Stale References": stale_in_case,
            "New Items": new_in_case,
            "Missing Previous Dates": missing_previous_in_case,
            "Missing Warranty Quotes": warranty_missing_in_case,
            "Risk Score": risk_score,
            "Risk Level": risk_level,
            "Risk Flags": "; ".join(risks),
        })

    distinct_case_count = len(distinct_cases)
    competitive_cases = distinct_case_count - single_source_cases
    competitive_pct = (competitive_cases / distinct_case_count * 100) if distinct_case_count else 0

    overall_risk_score = 0
    if duplicate_sheets:
        overall_risk_score += 15
    if single_source_cases:
        overall_risk_score += 20
    if stale_references:
        overall_risk_score += 15
    if missing_warranty_quotes:
        overall_risk_score += 10
    if extended_lead_time_quotes:
        overall_risk_score += 10
    if duplicate_vendor_cases:
        overall_risk_score += 15
    overall_risk_score = min(overall_risk_score, 100)

    if overall_risk_score >= 60:
        overall_risk_level = "Critical"
    elif overall_risk_score >= 40:
        overall_risk_level = "High"
    elif overall_risk_score >= 20:
        overall_risk_level = "Moderate"
    else:
        overall_risk_level = "Low"

    return {
        "analysis_type": "commercial_comparison",
        "overall": {
            "uploaded_sheets": len(cases),
            "distinct_procurement_cases": distinct_case_count,
            "duplicate_sheets": len(duplicate_sheets),
            "total_items": total_items,
            "vendor_quotation_columns": total_vendor_columns,
            "competitive_cases": competitive_cases,
            "competitive_case_pct": round(competitive_pct, 1),
            "single_source_cases": single_source_cases,
            "stale_references": stale_references,
            "recent_references": recent_references,
            "new_item_references": new_item_references,
            "missing_previous_dates": missing_previous_dates,
            "missing_warranty_quotes": missing_warranty_quotes,
            "extended_lead_time_quotes": extended_lead_time_quotes,
            "overall_risk_score": overall_risk_score,
            "overall_risk_level": overall_risk_level,
        },
        "case_rows": case_rows,
        "item_rows": item_rows,
        "vendor_rows": vendor_rows,
        "duplicate_rows": duplicate_sheets,
    }


def build_commercial_insights(result):
    overall = result["overall"]
    risk_level = f"{overall['overall_risk_level']} ({overall['overall_risk_score']}/100)"

    summary_points = [
        f"{overall['distinct_procurement_cases']} distinct procurement cases were analyzed from "
        f"{overall['uploaded_sheets']} uploaded sheets.",
        f"{overall['competitive_case_pct']:.1f}% of cases have competitive supplier coverage.",
        f"{overall['single_source_cases']} procurement case(s) have single-source risk.",
        f"{overall['stale_references']} previous-order reference(s) are older than 12 months.",
        f"{overall['missing_warranty_quotes']} supplier quotation(s) have missing warranty terms.",
        f"{overall['duplicate_sheets']} duplicate sheet(s) were excluded from distinct-case totals.",
    ]

    actions = []
    if overall["single_source_cases"]:
        actions.append("Obtain additional supplier quotations or record an approved single-source justification.")
    if overall["stale_references"]:
        actions.append("Refresh market quotations for previous-order references older than 12 months.")
    if overall["missing_warranty_quotes"]:
        actions.append("Confirm warranty duration, coverage, exclusions, and replacement conditions before PO release.")
    if overall["duplicate_sheets"]:
        actions.append("Remove duplicate procurement sheets or assign unique PR/RFQ identifiers.")

    actions.extend([
        "Add PR number, item code, item description, quantity, unit price, tax, freight amount, and "
        "total landed cost to future uploads.",
        "Capture promised delivery, actual receipt, accepted quantity, and rejected quantity for "
        "supplier performance measurement.",
    ])

    recommendation = (
        "Treat the current workbook as a sourcing and commercial-terms comparison rather than a "
        "financial performance report. Resolve duplicate records, single-source cases, stale "
        "benchmarks, and missing warranty information before supplier award. Add price, quantity, "
        "PO, receipt, and quality fields to enable spend, savings, delivery, and supplier-performance "
        "analysis."
    )

    return {
        "risk_level": risk_level,
        "summary_points": summary_points,
        "recommended_actions": actions,
        "recommendation": recommendation,
    }


def make_commercial_ai_prompt(result, insights=None):
    overall = result["overall"]
    insights = insights or build_commercial_insights(result)

    case_lines = [
        f"- {row['Sheet']}: {row['Distinct Vendors']} vendor(s), {row['Items']} item(s), "
        f"Risk {row['Risk Level']} ({row['Risk Score']}/100)"
        + (f" — {row['Risk Flags']}" if row["Risk Flags"] else "")
        for row in result["case_rows"]
    ]

    return f"""
PROCUREMENT COMMERCIAL COMPARISON — ANALYSIS CONTEXT

Overall Risk: {insights['risk_level']}

Overview:
- Uploaded Sheets: {overall['uploaded_sheets']}
- Distinct Procurement Cases: {overall['distinct_procurement_cases']}
- Duplicate Sheets: {overall['duplicate_sheets']}
- Total Item References: {overall['total_items']}
- Competitive Cases: {overall['competitive_case_pct']}%
- Single-Source Cases: {overall['single_source_cases']}
- Stale References: {overall['stale_references']}
- Missing Warranty Quotes: {overall['missing_warranty_quotes']}

Executive Summary:
{chr(10).join('- ' + p for p in insights['summary_points'])}

Recommended Actions:
{chr(10).join('- ' + a for a in insights['recommended_actions'])}

Recommendation:
{insights['recommendation']}

Case-Level Breakdown:
{chr(10).join(case_lines)}
""".strip()
