"""
Procurement / Vendor Price Comparison engine — Manufacturing industry module
for AI Operations Manager.

Parses "Price Comparative Sheet" style workbooks (one sheet per Purchase
Requisition, N vendor quote blocks per sheet, plus a Previous Order Details
block) and produces vendor-comparison analysis: recommended vendor per item,
savings vs alternatives, price trend vs the last order, and risk flags.

This is intentionally a SEPARATE module from engine.py (the BPO/ops-KPI
engine) — Manufacturing procurement analysis has nothing in common with
Productivity/Quality/SLA/AHT, so it doesn't touch or depend on engine.py.
"""

import re
from datetime import datetime

import openpyxl
import pandas as pd


def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def _is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _find_header_row(ws, max_scan=20):
    for r in range(1, max_scan + 1):
        val = _clean(ws.cell(row=r, column=1).value)
        if val == "S.N.":
            return r
    return None


def _find_vendor_row(ws, header_row):
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            if "previous order details" in _clean(ws.cell(row=r, column=c).value).lower():
                return r
    return header_row - 3 if header_row and header_row > 3 else None


def _parse_sheet(ws, sheet_name):
    header_row = _find_header_row(ws)
    if header_row is None:
        return None  # not a Price Comparative Sheet — skip silently

    vendor_row = _find_vendor_row(ws, header_row)
    max_col = ws.max_column

    headers = {c: _clean(ws.cell(row=header_row, column=c).value) for c in range(1, max_col + 1)}

    # Fixed leading columns
    col_sn, col_prno, col_itemcode, col_itemname, col_qty, col_um = 1, 2, 3, 4, 5, 6

    # Walk columns from 7 onward: consume 5-col vendor blocks (Make, Quoted
    # Price, Dis %, Dis Price, Total Price) until we hit "Doc. Date".
    vendor_blocks = []  # list of dicts: {name, make_col, quoted_col, disp_pct_col, disp_price_col, total_col}
    c = 7
    while c <= max_col:
        label = headers.get(c, "")
        if label == "Doc. Date":
            break
        if label.startswith("Make"):
            vendor_name = _clean(ws.cell(row=vendor_row, column=c).value) if vendor_row else f"Vendor@{c}"
            vendor_blocks.append({
                "name": vendor_name or f"Vendor@{c}",
                "make_col": c,
                "quoted_col": c + 1,
                "disc_pct_col": c + 2,
                "disc_price_col": c + 3,
                "total_col": c + 4,
            })
            c += 5
        else:
            c += 1

    # Previous order block: find Doc.Date, Vendor, DPTPL columns by label
    doc_date_col = vendor_col = dptpl_col = None
    for c2 in range(c, max_col + 1):
        label = headers.get(c2, "")
        if label == "Doc. Date":
            doc_date_col = c2
        elif label == "Vendor":
            vendor_col = c2
        elif label == "DPTPL":
            dptpl_col = c2

    # Meta info (plant/indenter/working date) — best-effort, from row 3/4 text
    plant_indenter = ""
    working_date = ""
    for r in range(1, header_row):
        text = _clean(ws.cell(row=r, column=1).value)
        if text.lower().startswith("plant"):
            plant_indenter = text
        if "working date" in text.lower():
            m = re.search(r"working date\s*:?\s*([0-9]{1,2}[-.][0-9]{1,2}[-.][0-9]{2,4})", text, re.IGNORECASE)
            if m:
                working_date = m.group(1)

    items = []
    r = header_row + 1
    while r <= ws.max_row:
        sn_val = ws.cell(row=r, column=col_sn).value
        if not _is_number(sn_val):
            break  # hit "Freight" row or end of item block

        item_code = _clean(ws.cell(row=r, column=col_itemcode).value)
        item_name = _clean(ws.cell(row=r, column=col_itemname).value)
        qty = ws.cell(row=r, column=col_qty).value
        um = _clean(ws.cell(row=r, column=col_um).value)
        pr_no = _clean(ws.cell(row=r, column=col_prno).value)

        quotes = []
        for vb in vendor_blocks:
            total = ws.cell(row=r, column=vb["total_col"]).value
            unit = ws.cell(row=r, column=vb["disc_price_col"]).value
            if _is_number(total) and total > 0:
                quotes.append({
                    "vendor": vb["name"],
                    "unit_price": unit if _is_number(unit) else None,
                    "total_price": total,
                })

        prev_price = ws.cell(row=r, column=dptpl_col).value if dptpl_col else None
        prev_vendor = _clean(ws.cell(row=r, column=vendor_col).value) if vendor_col else ""
        prev_date = _clean(ws.cell(row=r, column=doc_date_col).value) if doc_date_col else ""

        items.append({
            "pr_no": pr_no,
            "item_code": item_code,
            "item_name": item_name,
            "qty": qty,
            "um": um,
            "quotes": quotes,
            "prev_unit_price": prev_price if _is_number(prev_price) else None,
            "prev_vendor": prev_vendor,
            "prev_date": prev_date,
        })
        r += 1

    return {
        "sheet_name": sheet_name,
        "plant_indenter": plant_indenter,
        "working_date": working_date,
        "vendor_count": len(vendor_blocks),
        "items": items,
    }


def parse_workbook(filepath):
    """Parse every Price-Comparative-Sheet-style sheet in the workbook."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    prs = []
    for sn in wb.sheetnames:
        parsed = _parse_sheet(wb[sn], sn)
        if parsed and parsed["items"]:
            prs.append(parsed)
    return prs


def analyze_procurement(prs, price_increase_risk_pct=10.0):
    """
    Turn parsed PR sheets into item-level and PR-level analysis.
    Returns dict with 'item_rows' (list) and 'pr_rows' (list) and 'overall' (dict).
    """
    item_rows = []

    for pr in prs:
        pr_recommended_spend = 0.0
        pr_highest_spend = 0.0
        pr_savings = 0.0
        single_vendor_count = 0
        price_increase_count = 0
        no_prev_data_count = 0

        for it in pr["items"]:
            quotes = sorted(it["quotes"], key=lambda q: q["total_price"])
            n_quotes = len(quotes)

            if n_quotes == 0:
                # No valid quote at all — flag and skip ranking
                item_rows.append({
                    "PR Sheet": pr["sheet_name"],
                    "PR No": it["pr_no"],
                    "Item Code": it["item_code"],
                    "Item Name": it["item_name"],
                    "Qty": it["qty"],
                    "UM": it["um"],
                    "Vendors Quoted": 0,
                    "Recommended Vendor": None,
                    "Recommended Total Price": None,
                    "2nd Best Vendor": None,
                    "2nd Best Total Price": None,
                    "Highest Quoted Vendor": None,
                    "Highest Total Price": None,
                    "Savings vs Highest": None,
                    "Savings vs Highest %": None,
                    "Previous Unit Price": it["prev_unit_price"],
                    "Previous Vendor": it["prev_vendor"],
                    "% Change vs Previous": None,
                    "Risk Flags": "No vendor quote available",
                })
                continue

            best = quotes[0]
            worst = quotes[-1]
            second = quotes[1] if n_quotes > 1 else None

            savings_vs_highest = worst["total_price"] - best["total_price"]
            savings_vs_highest_pct = (
                (savings_vs_highest / worst["total_price"] * 100.0)
                if worst["total_price"] else None
            )

            pct_change_vs_prev = None
            if it["prev_unit_price"] and best["unit_price"] is not None and it["prev_unit_price"] > 0:
                pct_change_vs_prev = (
                    (best["unit_price"] - it["prev_unit_price"]) / it["prev_unit_price"] * 100.0
                )

            risk_flags = []
            if n_quotes == 1:
                risk_flags.append("Single vendor quoted")
                single_vendor_count += 1
            if pct_change_vs_prev is not None and pct_change_vs_prev >= price_increase_risk_pct:
                risk_flags.append(f"Price up {pct_change_vs_prev:.1f}% vs last order")
                price_increase_count += 1
            if it["prev_unit_price"] is None:
                no_prev_data_count += 1

            pr_recommended_spend += best["total_price"]
            pr_highest_spend += worst["total_price"]
            pr_savings += savings_vs_highest

            item_rows.append({
                "PR Sheet": pr["sheet_name"],
                "PR No": it["pr_no"],
                "Item Code": it["item_code"],
                "Item Name": it["item_name"],
                "Qty": it["qty"],
                "UM": it["um"],
                "Vendors Quoted": n_quotes,
                "Recommended Vendor": best["vendor"],
                "Recommended Total Price": round(best["total_price"], 2),
                "2nd Best Vendor": second["vendor"] if second else None,
                "2nd Best Total Price": round(second["total_price"], 2) if second else None,
                "Highest Quoted Vendor": worst["vendor"],
                "Highest Total Price": round(worst["total_price"], 2),
                "Savings vs Highest": round(savings_vs_highest, 2),
                "Savings vs Highest %": round(savings_vs_highest_pct, 1) if savings_vs_highest_pct is not None else None,
                "Previous Unit Price": it["prev_unit_price"],
                "Previous Vendor": it["prev_vendor"],
                "% Change vs Previous": round(pct_change_vs_prev, 1) if pct_change_vs_prev is not None else None,
                "Risk Flags": "; ".join(risk_flags) if risk_flags else "",
            })

        item_count = len(pr["items"])
        item_rows_for_pr = [r for r in item_rows if r["PR Sheet"] == pr["sheet_name"]]

    pr_rows = []
    for pr in prs:
        rows = [r for r in item_rows if r["PR Sheet"] == pr["sheet_name"]]
        recommended_spend = sum(r["Recommended Total Price"] or 0 for r in rows)
        highest_spend = sum(r["Highest Total Price"] or 0 for r in rows)
        savings = highest_spend - recommended_spend
        single_vendor_items = sum(1 for r in rows if "Single vendor" in (r["Risk Flags"] or ""))
        price_increase_items = sum(1 for r in rows if "Price up" in (r["Risk Flags"] or ""))
        no_quote_items = sum(1 for r in rows if r["Vendors Quoted"] == 0)

        pr_rows.append({
            "PR Sheet": pr["sheet_name"],
            "Plant / Indenter": pr["plant_indenter"],
            "Working Date": pr["working_date"],
            "Vendors Compared": pr["vendor_count"],
            "Items": len(rows),
            "Recommended Spend": round(recommended_spend, 2),
            "Highest-Quote Spend": round(highest_spend, 2),
            "Potential Savings": round(savings, 2),
            "Savings %": round(savings / highest_spend * 100.0, 1) if highest_spend else None,
            "Single-Vendor Items (Risk)": single_vendor_items,
            "Price-Increase Items (Risk)": price_increase_items,
            "No-Quote Items (Risk)": no_quote_items,
        })

    total_recommended = sum(r["Recommended Spend"] for r in pr_rows)
    total_highest = sum(r["Highest-Quote Spend"] for r in pr_rows)
    overall = {
        "total_prs": len(pr_rows),
        "total_items": sum(r["Items"] for r in pr_rows),
        "total_recommended_spend": round(total_recommended, 2),
        "total_highest_quote_spend": round(total_highest, 2),
        "total_potential_savings": round(total_highest - total_recommended, 2),
        "overall_savings_pct": round((total_highest - total_recommended) / total_highest * 100.0, 1) if total_highest else None,
        "total_single_vendor_items": sum(r["Single-Vendor Items (Risk)"] for r in pr_rows),
        "total_price_increase_items": sum(r["Price-Increase Items (Risk)"] for r in pr_rows),
        "total_no_quote_items": sum(r["No-Quote Items (Risk)"] for r in pr_rows),
    }

    return {"item_rows": item_rows, "pr_rows": pr_rows, "overall": overall}


# ============================================================
# INSIGHT LAYER
# Mirrors engine.py's analyze_data()-then-insight pattern (risk level,
# executive summary bullets, one management recommendation) so both
# industries produce the same KIND of insight from very different data.
# ============================================================

def compute_risk_level(overall):
    """Breach-based risk level — same 0/1/2/3+ pattern as the BPO engine,
    counting risk categories instead of KPI misses."""
    breaches = sum([
        overall["total_single_vendor_items"] > 0,
        overall["total_price_increase_items"] > 0,
        overall["total_no_quote_items"] > 0,
    ])
    level = {
        0: "🟢 LOW RISK",
        1: "🟡 MEDIUM RISK",
        2: "🟠 HIGH RISK",
    }.get(breaches, "🔴 CRITICAL RISK")
    return level, breaches


def build_summary_points(overall):
    """Plain-language bullets summarizing the procurement risk position."""
    return [
        f"{'🟢' if overall['total_potential_savings'] > 0 else '⚪'} "
        f"Potential savings: ₹{overall['total_potential_savings']:,.0f} "
        f"({overall['overall_savings_pct']}%) identified across "
        f"{overall['total_prs']} purchase requisition(s).",

        f"{'🔴' if overall['total_single_vendor_items'] > 0 else '🟢'} "
        f"Single-vendor items: {overall['total_single_vendor_items']} — "
        "no competitive comparison available for these purchases.",

        f"{'🟠' if overall['total_price_increase_items'] > 0 else '🟢'} "
        f"Price-increase items: {overall['total_price_increase_items']} — "
        "priced 10%+ higher than the last recorded order.",

        f"{'🔴' if overall['total_no_quote_items'] > 0 else '🟢'} "
        f"No-quote items: {overall['total_no_quote_items']} — no valid "
        "vendor quote was received.",
    ]


def build_recommendation(breaches):
    """One management recommendation, tone-matched to the BPO engine's."""
    if breaches >= 3:
        return (
            "Immediate procurement review is recommended. Multiple risk "
            "categories are present across this batch — prioritize sourcing "
            "additional vendor quotes and validating the flagged price "
            "increases before approving these purchase requisitions."
        )
    if breaches >= 1:
        return (
            "Procurement should review the flagged items, confirm pricing "
            "with the recommended vendors, and source a second quote for "
            "any single-vendor items where feasible."
        )
    return (
        "All items have competitive quotes with no unusual price movement. "
        "Proceed with the recommended vendors listed in the Item Detail report."
    )


def build_insights(result):
    """
    One call that returns everything the UI/report needs: risk level,
    summary bullets, and the recommendation — computed once so app.py,
    the Excel report, and the AI Copilot context never disagree.
    """
    overall = result["overall"]
    risk_level, breaches = compute_risk_level(overall)
    summary_points = build_summary_points(overall)
    recommendation = build_recommendation(breaches)
    return {
        "risk_level": risk_level,
        "breaches": breaches,
        "summary_points": summary_points,
        "recommendation": recommendation,
    }


def make_ai_prompt(result, insights=None):
    """
    Procurement equivalent of engine.py's make_ai_prompt() — a single
    text block an LLM (or the debug expander) can use as full context.
    """
    overall = result["overall"]
    insights = insights or build_insights(result)

    item_lines = []
    for row in result["item_rows"][:200]:  # keep the prompt bounded
        item_lines.append(
            f"- {row['Item Name']} (PR {row['PR Sheet']}, Qty {row['Qty']} {row['UM']}): "
            f"recommended {row['Recommended Vendor']} at ₹{row['Recommended Total Price']}, "
            f"highest quote ₹{row['Highest Total Price']}"
            + (f", 2nd best {row['2nd Best Vendor']} at ₹{row['2nd Best Total Price']}" if row['2nd Best Vendor'] else "")
            + (f", prev. unit price ₹{row['Previous Unit Price']}" if row['Previous Unit Price'] is not None else "")
            + (f", risk: {row['Risk Flags']}" if row['Risk Flags'] else "")
        )

    pr_lines = [
        f"- {row['PR Sheet']} ({row['Plant / Indenter'] or 'plant n/a'}, {row['Working Date'] or 'date n/a'}): "
        f"{row['Items']} items, {row['Vendors Compared']} vendors compared, "
        f"₹{row['Recommended Spend']:,.0f} recommended spend, "
        f"₹{row['Potential Savings']:,.0f} potential savings"
        for row in result["pr_rows"]
    ]

    return f"""
PROCUREMENT PRICE COMPARISON — ANALYSIS CONTEXT

Overall Risk: {insights['risk_level']}

Overview:
- Purchase Requisitions: {overall['total_prs']}
- Total Items: {overall['total_items']}
- Recommended Spend: ₹{overall['total_recommended_spend']:,.0f}
- Highest-Quote Spend: ₹{overall['total_highest_quote_spend']:,.0f}
- Potential Savings: ₹{overall['total_potential_savings']:,.0f} ({overall['overall_savings_pct']}%)
- Single-Vendor Items: {overall['total_single_vendor_items']}
- Price-Increase Items: {overall['total_price_increase_items']}
- No-Quote Items: {overall['total_no_quote_items']}

Executive Summary:
{chr(10).join('- ' + p for p in insights['summary_points'])}

Recommendation:
{insights['recommendation']}

Purchase Requisition Breakdown:
{chr(10).join(pr_lines)}

Item-Level Detail:
{chr(10).join(item_lines)}
""".strip()
