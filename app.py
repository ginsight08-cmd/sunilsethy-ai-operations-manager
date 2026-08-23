import io
import time
import json
import smtplib
from datetime import datetime, timedelta
from pathlib import Path
from email.message import EmailMessage

import pandas as pd
import requests
import streamlit as st
from supabase import create_client, Client

from engine import analyze_data, make_ai_prompt
import procurement_engine


# ============================================================
# GENERATIVE INSIGHT | AI OPERATIONS COPILOT
# Complete Streamlit application
# ============================================================

APP_NAME = "Generative Insight"
APP_VERSION = "1.0.0"

st.set_page_config(
    page_title="Generative Insight | AI Operations Copilot",
    page_icon="assests/Generative_insight.png",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ============================================================
# BRAND CONFIGURATION
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

# Add your logo here:
# assets/Generative_insight.png
LOGO_PATH = BASE_DIR / "assets" / "Generative_insight.png"

WEBSITE_URL = "https://generativeinsight.in"

BRAND_BLUE = "#0757B8"
BRAND_CYAN = "#00AEEF"
BRAND_ORANGE = "#FF9D00"
BRAND_NAVY = "#071A3D"


# ============================================================
# SESSION STATE
# ============================================================

DEFAULT_STATE = {
    "authenticated": False,
    "user_email": "",
    "user_id": "",
    "user_name": "",
    "company_name": "",
    "user_plan": "Free",
    "n8n_sent": False,
    "n8n_result": None,
    "copilot_answer": None,
    "last_question": "",
    "file_name": "",
    "analysis_result": None,
    "analysis_df": None,
    "report_pdf": None,
    "report_generated_at": None,
    "show_plans": False,
    "razorpay_checkout_url": "",
    "razorpay_subscription_id": "",
    "account_created_at": "",
    # Free-tier (₹299/mo, post 15-day trial) billing — tracked
    # separately from the Professional Razorpay fields above so the
    # two subscriptions never collide.
    "free_billing_status": "",
    "free_subscription_id": "",
    "razorpay_checkout_url_free": "",
    # Industry — switchable anytime from the sidebar. "BPO" uses the
    # existing Productivity/Quality/SLA/AHT engine; "Manufacturing" uses
    # the separate procurement/vendor-comparison engine below.
    "industry": "BPO",
    "manufacturing_file_name": "",
    "manufacturing_result": None,
    "manufacturing_report_bytes": None,
    "manufacturing_report_generated_at": None,
}

# Free plan trial length. After this many days on the Free plan,
# the dashboard is locked and the user is shown an upgrade-only screen.
FREE_TRIAL_DAYS = 15

# ============================================================
# INDUSTRY REGISTRY
# Single source of truth for the sidebar dropdown. "built" industries
# have a real, complete flow (their own upload, engine, tabs, insights).
# Everything else in this dict shows a graceful "coming soon" screen
# instead of crashing or showing fake data — add a new engine module
# (like procurement_engine.py) and flip it into BUILT_INDUSTRIES when
# a module is ready.
# ============================================================

INDUSTRY_LABELS = {
    "BPO": "BPO / Call Center Operations",
    "Manufacturing": "Manufacturing (Procurement)",
    "Retail": "Retail / E-commerce",
    "Logistics": "Logistics / Delivery",
    "Healthcare": "Healthcare / Clinics",
}

BUILT_INDUSTRIES = {"BPO", "Manufacturing"}

for key, value in DEFAULT_STATE.items():
    if key not in st.session_state:
        st.session_state[key] = value


# ============================================================
# BRAND THEME
# ============================================================

st.markdown(
    f"""
<style>

    /* ============================================================
       DESIGN TOKENS — matched to the approved mockups: minimal
       black/white surface, green for "good"/Free, amber for trial
       urgency, red for risk, blue reserved for links/the "Insight"
       wordmark accent. Card shapes, spacing, and button style below
       all follow the mockups; every widget/function is unchanged.
       ============================================================ */
    :root {{
        --ink: #0F1115;
        --ink-soft: #4B5262;
        --muted: #6B7280;
        --paper: #FFFFFF;
        --bg: #FAFAFA;
        --border: #E7E8EC;
        --border-soft: #EFEFF2;
        --black: #111114;
        --black-hover: #26272C;
        --green-bg: #E8F8EE;
        --green-ink: #1D8A4A;
        --amber-bg: #FDF3DA;
        --amber-ink: #8A6100;
        --red-bg: #FDEBEC;
        --red-ink: #B42318;
        --blue-link: {BRAND_BLUE};
        --radius-lg: 20px;
        --radius-md: 14px;
        --radius-sm: 10px;
        --shadow-card: 0 1px 2px rgba(15,17,21,0.04), 0 8px 20px rgba(15,17,21,0.04);
    }}

    .stApp {{
        background: var(--bg);
    }}

    .main {{
        padding-top: 1.25rem;
    }}

    /* ---------- Typography ---------- */
    .main-title {{
        font-size: 2.1rem;
        font-weight: 800;
        color: var(--ink);
        margin-bottom: 0.2rem;
        letter-spacing: -0.02em;
    }}

    .brand-subtitle {{
        color: var(--muted) !important;
        -webkit-text-fill-color: var(--muted) !important;
        font-size: 1rem;
        font-weight: 500;
        margin-bottom: 1rem;
        opacity: 1 !important;
        visibility: visible !important;
    }}

    /* ---------- Logo mark (CSS-drawn, matches the mockup's black
       rounded-square sparkle icon; a real uploaded logo image always
       takes priority over this — see show_brand_header()). ---------- */
    .gi-logo-mark {{
        width: 40px; height: 40px; border-radius: 11px;
        background: var(--black);
        display: inline-flex; align-items: center; justify-content: center;
        font-size: 18px; color: #FFFFFF; flex-shrink: 0;
        box-shadow: 0 2px 6px rgba(17,17,20,0.25);
    }}
    .gi-brand-row {{ display: flex; align-items: center; gap: 12px; margin-bottom: 2px; }}

    /* ---------- Auth screen — matched to the provided reference
       design (Tailwind neutral-950/500/200 palette, rounded-2xl
       card, size-14 icon badge). Same form fields/logic underneath,
       purely presentational wrapper. ---------- */
    .gi-auth-topbar {{
        display: flex; align-items: center; justify-content: space-between;
        padding: 18px 8px; border-bottom: 1px solid #E5E5E5;
        margin-bottom: 8px;
    }}
    .gi-auth-icon-sm {{
        width: 40px; height: 40px; border-radius: 12px;
        background: #0A0A0A; color: #FFFFFF;
        display: flex; align-items: center; justify-content: center;
        font-size: 18px; flex-shrink: 0;
    }}
    .gi-auth-icon-lg {{
        width: 56px; height: 56px; border-radius: 16px;
        background: #0A0A0A; color: #FFFFFF;
        display: flex; align-items: center; justify-content: center;
        font-size: 26px; margin: 0 auto 14px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    .gi-auth-hero {{ text-align: center; margin-bottom: 4px; }}
    .gi-auth-hero-title {{
        font-size: 1.85rem; font-weight: 700; letter-spacing: -0.02em;
        color: #0A0A0A; margin-bottom: 6px;
    }}
    .gi-auth-hero-sub {{
        font-size: 0.9rem; color: #737373; max-width: 380px;
        margin: 0 auto;
    }}
    .gi-auth-card-title {{
        font-size: 1.25rem; font-weight: 700; color: #0A0A0A;
        margin-bottom: 2px;
    }}
    .gi-auth-card-desc {{
        font-size: 0.88rem; color: #737373; margin-bottom: 18px;
    }}
    .gi-field-label-row {{
        display: flex; align-items: center; gap: 7px;
        font-size: 0.85rem; font-weight: 600; color: #0A0A0A;
        margin-bottom: 4px; margin-top: 2px;
    }}
    .gi-auth-footer-note {{
        text-align: center; font-size: 0.78rem; color: #737373;
        margin-top: 18px; line-height: 1.5;
    }}
    .gi-auth-shell {{
        max-width: 460px; margin: 8px auto 0;
        background: #FFFFFF; border: 1px solid #E5E5E5;
        border-radius: 16px; padding: 2rem;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }}

    .gi-brand {{
        font-size: 1.3rem;
        font-weight: 800;
        color: var(--ink);
        letter-spacing: -0.02em;
        line-height: 1.15;
    }}

    .gi-brand span {{
        color: var(--blue-link);
    }}

    .gi-tagline {{
        color: var(--muted);
        font-size: 0.78rem;
        margin-top: 0.1rem;
    }}

    /* ---------- Hero / info cards ---------- */
    .hero {{
        padding: 1.4rem 1.6rem;
        border-radius: var(--radius-lg);
        border: 1px solid var(--border);
        background: var(--paper);
        box-shadow: var(--shadow-card);
        margin-bottom: 1.2rem;
    }}

    .small-muted {{
        color: var(--muted);
        font-size: .88rem;
    }}

    /* ---------- st.metric cards ---------- */
    div[data-testid="stMetric"] {{
        background: var(--paper);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        padding: 1rem 1.1rem;
        box-shadow: var(--shadow-card);
    }}

    div[data-testid="stMetricLabel"] {{
        color: var(--muted) !important;
        font-weight: 600 !important;
    }}

    div[data-testid="stMetricValue"] {{
        color: var(--ink) !important;
        font-weight: 800 !important;
    }}

    div[data-testid="stMetricDelta"] svg {{
        display: inline;
    }}

    /* ---------- Custom metric-card component (used via
       render_metric_card()) — icon + label + big value + colored
       delta chip, matching the mockup's stat cards more closely
       than st.metric's default layout allows. ---------- */
    .gi-metric-card {{
        background: var(--paper);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        padding: 1.1rem 1.2rem;
        box-shadow: var(--shadow-card);
        height: 100%;
    }}
    .gi-metric-icon {{
        font-size: 1.05rem;
        margin-bottom: 0.55rem;
        opacity: 0.85;
    }}
    .gi-metric-label {{
        color: var(--muted);
        font-size: 0.82rem;
        font-weight: 600;
        margin-bottom: 0.3rem;
    }}
    .gi-metric-value {{
        color: var(--ink);
        font-size: 1.55rem;
        font-weight: 800;
        letter-spacing: -0.01em;
        line-height: 1.1;
        margin-bottom: 0.4rem;
    }}
    .gi-metric-delta {{
        display: inline-flex;
        align-items: center;
        gap: 4px;
        font-size: 0.78rem;
        font-weight: 700;
        padding: 2px 9px;
        border-radius: 999px;
    }}
    .gi-metric-delta.good {{ background: var(--green-bg); color: var(--green-ink); }}
    .gi-metric-delta.bad {{ background: var(--red-bg); color: var(--red-ink); }}
    .gi-metric-delta.neutral {{ background: var(--border-soft); color: var(--ink-soft); }}

    /* ---------- Pricing / plan cards ---------- */
    .plan-card {{
        padding: 1.35rem;
        border-radius: var(--radius-lg);
        border: 1px solid var(--border);
        background: var(--paper);
        min-height: 260px;
        box-shadow: var(--shadow-card);
        color: var(--ink) !important;
        transition: transform .18s ease, box-shadow .18s ease, border-color .18s ease;
        position: relative;
    }}

    .plan-card:hover {{
        border-color: var(--ink);
        box-shadow: 0 4px 8px rgba(15,17,21,0.06), 0 16px 32px rgba(15,17,21,0.08);
        transform: translateY(-3px);
    }}

    .plan-card .plan-badge-popular {{
        position: absolute; top: -11px; right: 18px;
        background: var(--black); color: #FFFFFF;
        font-size: 0.68rem; font-weight: 700;
        padding: 4px 12px; border-radius: 999px;
        letter-spacing: 0.02em;
    }}

    .plan-card .plan-name, .plan-card .plan-price {{
        color: var(--ink) !important;
    }}
    .plan-card .plan-description, .plan-card .plan-feature {{
        color: var(--muted) !important;
    }}
    .plan-card .plan-feature {{
        font-weight: 500 !important;
    }}

    /* ---------- Sidebar ---------- */
    section[data-testid="stSidebar"] {{
        background: var(--paper);
        border-right: 1px solid var(--border);
    }}

    /* Status pills used in the sidebar (plan badge, trial notice) */
    .gi-pill {{
        display: flex;
        align-items: flex-start;
        gap: 8px;
        border-radius: var(--radius-sm);
        padding: 10px 12px;
        font-size: 0.85rem;
        font-weight: 600;
        line-height: 1.4;
        margin-bottom: 10px;
    }}
    .gi-pill.green {{ background: var(--green-bg); color: var(--green-ink); }}
    .gi-pill.amber {{ background: var(--amber-bg); color: var(--amber-ink); }}
    .gi-pill.red {{ background: var(--red-bg); color: var(--red-ink); }}
    .gi-pill.neutral {{ background: var(--border-soft); color: var(--ink-soft); }}

    /* ---------- Buttons ---------- */
    .stButton > button {{
        border-radius: var(--radius-sm);
        font-weight: 700;
        border: 1px solid var(--border);
        color: var(--ink);
        background: var(--paper);
    }}

    .stButton > button:hover {{
        border-color: var(--ink);
    }}

    .stButton > button[kind="primary"] {{
        background: var(--black);
        color: #FFFFFF;
        border: none;
    }}

    .stButton > button[kind="primary"]:hover {{
        background: var(--black-hover);
        color: #FFFFFF;
    }}

    [data-testid="stLinkButton"] a {{
        border-radius: var(--radius-sm) !important;
        font-weight: 700 !important;
    }}

    /* ---------- Tabs ---------- */
    button[data-baseweb="tab"] {{
        font-weight: 700;
    }}

    button[data-baseweb="tab"][aria-selected="true"] {{
        color: var(--ink);
    }}

    .stApp [data-baseweb="tab-list"] {{
        background: transparent !important;
        border-bottom: 1px solid var(--border) !important;
    }}
    .stApp button[data-baseweb="tab"],
    .stApp button[data-baseweb="tab"] span,
    .stApp button[data-baseweb="tab"] div {{
        color: var(--muted) !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: var(--muted) !important;
        font-weight: 700 !important;
    }}
    .stApp button[data-baseweb="tab"][aria-selected="true"],
    .stApp button[data-baseweb="tab"][aria-selected="true"] span,
    .stApp button[data-baseweb="tab"][aria-selected="true"] div {{
        color: var(--ink) !important;
        -webkit-text-fill-color: var(--ink) !important;
    }}
    .stApp [data-baseweb="tab-highlight"] {{
        background-color: var(--ink) !important;
    }}

    a {{
        color: var(--blue-link);
    }}
    .stApp a {{
        color: var(--blue-link) !important;
        text-decoration: none;
    }}
    .stApp a:hover {{
        text-decoration: underline;
    }}

    .gi-footer {{
        text-align: center;
        color: var(--muted);
        font-size: 0.82rem;
        padding: 1.5rem 0;
    }}

    /* ============================================================
       EMBED-SAFE TYPOGRAPHY — explicit dark text prevents host-page
       CSS from making Streamlit content white/invisible in an iframe.
       ============================================================ */
    .stApp .stMarkdown, .stApp .stMarkdown p, .stApp .stMarkdown li,
    .stApp .stMarkdown span, .stApp label,
    .stApp [data-testid="stWidgetLabel"], .stApp [data-testid="stWidgetLabel"] * {{
        color: var(--ink) !important;
    }}
    .stApp .stCaption, .stApp [data-testid="stCaptionContainer"],
    .stApp [data-testid="stCaptionContainer"] * {{
        color: var(--muted) !important;
    }}
    .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6 {{
        color: var(--ink) !important;
    }}
    .stApp input, .stApp textarea, .stApp select,
    .stApp input::placeholder, .stApp textarea::placeholder {{
        color: var(--ink) !important;
        -webkit-text-fill-color: var(--ink) !important;
        opacity: 1 !important;
    }}
    .stApp [data-baseweb="select"] * {{
        color: var(--ink) !important;
    }}
    .stApp [data-baseweb="select"] > div {{
        border-radius: var(--radius-sm) !important;
        border-color: var(--border) !important;
    }}
    .stApp [data-testid="stDataFrame"] *,
    .stApp [data-testid="stTable"] * {{
        color: var(--ink) !important;
    }}
    .plan-card, .plan-card * {{
        color: var(--ink) !important;
    }}
    .stApp [data-testid="stMarkdownContainer"],
    .stApp [data-testid="stMarkdownContainer"] p,
    .stApp [data-testid="stMarkdownContainer"] div,
    .stApp [data-testid="stMarkdownContainer"] span,
    .stApp [data-testid="stMarkdownContainer"] li,
    .stApp [data-testid="stMarkdownContainer"] strong,
    .stApp [data-testid="stMarkdownContainer"] em {{
        color: var(--ink) !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: var(--ink) !important;
    }}
    .stApp [data-testid="stMarkdownContainer"] h1,
    .stApp [data-testid="stMarkdownContainer"] h2,
    .stApp [data-testid="stMarkdownContainer"] h3,
    .stApp [data-testid="stMarkdownContainer"] h4,
    .stApp [data-testid="stMarkdownContainer"] h5,
    .stApp [data-testid="stMarkdownContainer"] h6 {{
        color: var(--ink) !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: var(--ink) !important;
    }}
    .stApp .hero p {{
        color: var(--ink-soft) !important;
        opacity: 1 !important;
        visibility: visible !important;
        -webkit-text-fill-color: var(--ink-soft) !important;
        font-size: 0.96rem !important;
        line-height: 1.65 !important;
    }}
    .stApp .hero .brand-subtitle {{
        color: var(--blue-link) !important;
        opacity: 1 !important;
        -webkit-text-fill-color: var(--blue-link) !important;
    }}
    .stApp [data-testid="stHeader"] *,
    .stApp [data-testid="stText"],
    .stApp [data-testid="stCaptionContainer"] *,
    .stApp [data-testid="stWidgetLabel"] *,
    .stApp [data-testid="stForm"] label,
    .stApp [data-testid="stForm"] p {{
        opacity: 1 !important;
        visibility: visible !important;
    }}
    .stApp input,
    .stApp textarea,
    .stApp [role="textbox"],
    .stApp [data-baseweb="input"] input,
    .stApp [data-baseweb="textarea"] textarea {{
        background-color: var(--paper) !important;
        color: var(--ink) !important;
        -webkit-text-fill-color: var(--ink) !important;
        opacity: 1 !important;
    }}
    .stApp [data-baseweb="input"], .stApp [data-baseweb="textarea"] {{
        border-radius: var(--radius-sm) !important;
        border-color: var(--border) !important;
    }}
    .stApp input::placeholder,
    .stApp textarea::placeholder {{
        color: #9CA3AF !important;
        -webkit-text-fill-color: #9CA3AF !important;
        opacity: 1 !important;
    }}
    .stApp .stButton button,
    .stApp [data-testid="stFormSubmitButton"] button,
    .stApp [data-testid="stLinkButton"] a {{
        opacity: 1 !important;
        visibility: visible !important;
    }}
    .stApp .gi-brand, .stApp .gi-brand * {{ color:var(--ink) !important; -webkit-text-fill-color:var(--ink) !important; opacity:1 !important; visibility:visible !important; }}
    .stApp .gi-brand span {{ color:var(--blue-link) !important; -webkit-text-fill-color:var(--blue-link) !important; }}
    .stApp .gi-tagline {{ color:var(--muted) !important; -webkit-text-fill-color:var(--muted) !important; opacity:1 !important; }}
    .stApp .website-link {{ color:var(--blue-link) !important; -webkit-text-fill-color:var(--blue-link) !important; font-weight:700 !important; text-decoration:none !important; }}
    .stApp .plan-card, .stApp .plan-card * {{ opacity:1 !important; visibility:visible !important; }}

    /* ============================================================
       HIDE STREAMLIT CHROME
       ============================================================ */
    #MainMenu {{ visibility: hidden !important; }}
    header [data-testid="stToolbar"] {{ display: none !important; visibility: hidden !important; }}
    footer {{ visibility: hidden !important; }}
    [data-testid="stDecoration"] {{ display: none !important; }}
    a[href*="github.com"] {{ display: none !important; }}

    /* ============================================================
       RESPONSIVE — tablet (\u2264768px) and phone (\u2264480px). Every
       card/button/grid below reflows to single-column with full-width,
       touch-sized (\u226544px) controls at phone width.
       ============================================================ */
    @media (max-width: 768px) {{
        .main-title {{ font-size: 1.6rem !important; }}
        .hero {{ padding: 1.1rem 1rem !important; border-radius: 16px !important; }}
        .plan-card {{ min-height: auto !important; margin-bottom: .75rem !important; }}
        .gi-metric-card {{ margin-bottom: .6rem; }}
        .stApp .stButton > button,
        .stApp [data-testid="stLinkButton"] a,
        .stApp [data-testid="stFormSubmitButton"] button {{
            width: 100% !important;
            min-height: 44px !important;
        }}
        div[data-testid="column"] {{
            width: 100% !important;
            flex: 1 1 100% !important;
            min-width: 100% !important;
        }}
    }}

    @media (max-width: 480px) {{
        .main-title {{ font-size: 1.35rem !important; }}
        .brand-subtitle {{ font-size: 0.88rem !important; }}
        .hero {{ padding: 0.9rem 0.85rem !important; border-radius: 14px !important; }}
        .gi-metric-value {{ font-size: 1.3rem !important; }}
        div[data-testid="stMetricValue"] {{ font-size: 1.35rem !important; }}
        .plan-card {{ padding: 1.05rem !important; }}
        section[data-testid="stSidebar"] {{ min-width: 100% !important; }}
    }}
</style>
""",
    unsafe_allow_html=True,
)



# ============================================================
# HELPERS
# ============================================================

def render_metric_card(icon, label, value, delta=None, delta_tone="neutral"):
    """
    Custom stat card: icon, label, big value, colored delta chip.
    Visually closer to the approved mockups than st.metric's default
    layout — purely presentational, carries no logic of its own.
    delta_tone: "good" (green), "bad" (red), or "neutral" (gray).
    """
    delta_html = ""
    if delta:
        delta_html = f'<div class="gi-metric-delta {delta_tone}">{delta}</div>'
    st.markdown(
        f"""<div class="gi-metric-card">
<div class="gi-metric-icon">{icon}</div>
<div class="gi-metric-label">{label}</div>
<div class="gi-metric-value">{value}</div>
{delta_html}
</div>""",
        unsafe_allow_html=True,
    )


def render_user_badge(name, email):
    """Circular initials avatar + name/email, replacing plain-text caption lines."""
    label = (name or email or "?").strip()
    initial = label[0].upper() if label else "?"
    display_name = name.strip() if name and name.strip() else email
    st.markdown(
        f"""<div style="display:flex; align-items:center; gap:10px; margin:4px 0 2px;">
<div style="
    width:36px; height:36px; border-radius:50%;
    background:linear-gradient(135deg, {BRAND_BLUE}, {BRAND_CYAN});
    color:#FFFFFF; font-weight:800; font-size:15px;
    display:flex; align-items:center; justify-content:center;
    flex-shrink:0;
">{initial}</div>
<div style="min-width:0;">
<div style="font-weight:700; font-size:0.92rem; color:{BRAND_NAVY}; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">{display_name}</div>
<div style="font-size:0.78rem; color:#667085; white-space:nowrap; overflow:hidden; text-overflow:ellipsis;">{email}</div>
</div>
</div>""",
        unsafe_allow_html=True,
    )


def secret(name, default=""):
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


def show_brand_header(compact=False):
    """Display the Generative Insight logo and website branding."""

    if LOGO_PATH.exists():
        st.image(
            str(LOGO_PATH),
            width=230 if compact else 420,
        )
    else:
        st.markdown(
            """<div class="gi-brand-row">
<div class="gi-logo-mark">✦</div>
<div>
<div class="gi-brand">Generative <span>Insight</span></div>
<div class="gi-tagline">Insights today. Intelligence tomorrow.</div>
</div>
</div>""",
            unsafe_allow_html=True,
        )

    st.markdown(
        f"""<div style="margin-top:-8px; margin-bottom:18px; color:#667085; font-size:0.85rem;">
AI / ML &nbsp; | &nbsp; Annotation &nbsp; | &nbsp; Web & App Development
&nbsp;&nbsp;·&nbsp;&nbsp;
<a class="website-link" href="{WEBSITE_URL}" target="_blank" rel="noopener noreferrer">Visit Website</a>
</div>""",
        unsafe_allow_html=True,
    )


def get_supabase_client() -> Client:
    """Create the Supabase client from Streamlit Secrets."""
    url = secret("SUPABASE_URL")
    anon_key = secret("SUPABASE_ANON_KEY")

    if not url or not anon_key:
        raise RuntimeError(
            "Supabase authentication is not configured. "
            "Add SUPABASE_URL and SUPABASE_ANON_KEY to Streamlit Secrets."
        )

    return create_client(url, anon_key)


def get_supabase_admin_client() -> Client:
    url = secret("SUPABASE_URL")
    service_role_key = secret("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not service_role_key:
        raise RuntimeError("Add SUPABASE_SERVICE_ROLE_KEY to Streamlit Secrets for secure plan activation.")
    return create_client(url, service_role_key)


def update_user_plan(plan, subscription_id="", razorpay_status=""):
    if not st.session_state.get("user_id"):
        raise RuntimeError("No authenticated user is available.")
    admin = get_supabase_admin_client()
    current = admin.auth.admin.get_user_by_id(st.session_state.user_id)
    user = getattr(current, "user", None)
    metadata = dict(getattr(user, "user_metadata", {}) or {}) if user else {}
    metadata.update({
        "plan": plan,
        "razorpay_subscription_id": subscription_id or metadata.get("razorpay_subscription_id", ""),
        "razorpay_subscription_status": razorpay_status or metadata.get("razorpay_subscription_status", ""),
        "plan_updated_at": datetime.utcnow().isoformat() + "Z",
    })
    admin.auth.admin.update_user_by_id(st.session_state.user_id, {"user_metadata": metadata})
    st.session_state.user_plan = plan
    st.session_state.razorpay_subscription_id = metadata.get("razorpay_subscription_id", "")


def update_free_billing_status(status, subscription_id=""):
    """
    Tracks the ₹299/mo Free-tier subscription (post 15-day trial) WITHOUT
    changing the user's 'plan' field — someone on this billing plan is
    still logically on 'Free' feature limits, just paying to keep access
    past the trial window.
    """
    if not st.session_state.get("user_id"):
        raise RuntimeError("No authenticated user is available.")
    admin = get_supabase_admin_client()
    current = admin.auth.admin.get_user_by_id(st.session_state.user_id)
    user = getattr(current, "user", None)
    metadata = dict(getattr(user, "user_metadata", {}) or {}) if user else {}
    metadata.update({
        "free_billing_status": status,
        "free_subscription_id": subscription_id or metadata.get("free_subscription_id", ""),
        "free_billing_updated_at": datetime.utcnow().isoformat() + "Z",
    })
    admin.auth.admin.update_user_by_id(st.session_state.user_id, {"user_metadata": metadata})
    st.session_state.free_billing_status = status
    st.session_state.free_subscription_id = metadata.get("free_subscription_id", "")


def free_billing_is_active():
    """True once the ₹299/mo Free-tier subscription is confirmed active."""
    return st.session_state.get("free_billing_status", "") == "active"


def update_user_industry(industry):
    """Persist the selected industry (BPO / Manufacturing) to Supabase."""
    if not st.session_state.get("user_id"):
        st.session_state.industry = industry
        return
    try:
        admin = get_supabase_admin_client()
        current = admin.auth.admin.get_user_by_id(st.session_state.user_id)
        user = getattr(current, "user", None)
        metadata = dict(getattr(user, "user_metadata", {}) or {}) if user else {}
        metadata["industry"] = industry
        admin.auth.admin.update_user_by_id(st.session_state.user_id, {"user_metadata": metadata})
    except Exception:
        pass  # Non-critical — session state still reflects the choice this session.
    st.session_state.industry = industry


def get_razorpay_subscription(subscription_id):
    key_id, key_secret = secret("RAZORPAY_KEY_ID"), secret("RAZORPAY_KEY_SECRET")
    if not key_id or not key_secret:
        raise RuntimeError("Razorpay credentials are not configured.")
    if not subscription_id:
        raise RuntimeError("No Razorpay subscription ID is available.")
    try:
        response = requests.get(f"{RAZORPAY_API_BASE}/subscriptions/{subscription_id}", auth=(key_id, key_secret), timeout=30)
    except requests.exceptions.Timeout as exc:
        raise RuntimeError("Razorpay verification timed out. Please try again.") from exc
    except requests.exceptions.RequestException as exc:
        raise RuntimeError(f"Could not connect to Razorpay: {exc}") from exc
    try:
        data = response.json()
    except ValueError:
        data = {"error": response.text}
    if response.status_code >= 300:
        error = data.get("error", data) if isinstance(data, dict) else data
        if isinstance(error, dict):
            error = error.get("description") or error.get("reason") or str(error)
        raise RuntimeError(f"Razorpay verification failed (HTTP {response.status_code}): {error}")
    return data


def razorpay_activation_ready(plan_id_secret="RAZORPAY_PROFESSIONAL_PLAN_ID", plan_label="Professional"):
    if not st.session_state.get("authenticated") or not st.session_state.get("user_id"):
        return False, f"Please create an account or sign in before starting a {plan_label} subscription."
    if not secret("SUPABASE_SERVICE_ROLE_KEY"):
        return False, "Secure plan activation is not configured. Add SUPABASE_SERVICE_ROLE_KEY to Streamlit Secrets."
    if not razorpay_is_configured(plan_id_secret):
        return False, f"Razorpay is not fully configured. Add RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET and {plan_id_secret} to Streamlit Secrets."
    return True, ""


def verify_professional_subscription():
    ready, message = razorpay_activation_ready("RAZORPAY_PROFESSIONAL_PLAN_ID", "Professional")
    if not ready:
        raise RuntimeError(message)
    subscription_id = st.session_state.get("razorpay_subscription_id", "")
    if not subscription_id:
        admin = get_supabase_admin_client()
        current = admin.auth.admin.get_user_by_id(st.session_state.user_id)
        user = getattr(current, "user", None)
        metadata = getattr(user, "user_metadata", {}) or {} if user else {}
        subscription_id = metadata.get("razorpay_subscription_id", "")
    if not subscription_id:
        raise RuntimeError("No Razorpay subscription ID is available. Create the Professional checkout first.")
    data = get_razorpay_subscription(subscription_id)
    status = str(data.get("status", "")).lower()
    if status == "active":
        update_user_plan("Professional", subscription_id, status)
        return True, status
    update_user_plan("Free", subscription_id, status)
    return False, status


def verify_free_billing_subscription():
    """Verify the ₹299/mo Free-tier subscription. Does NOT change 'plan'."""
    ready, message = razorpay_activation_ready("RAZORPAY_FREE_PLAN_ID", "Free")
    if not ready:
        raise RuntimeError(message)
    subscription_id = st.session_state.get("free_subscription_id", "")
    if not subscription_id:
        admin = get_supabase_admin_client()
        current = admin.auth.admin.get_user_by_id(st.session_state.user_id)
        user = getattr(current, "user", None)
        metadata = getattr(user, "user_metadata", {}) or {} if user else {}
        subscription_id = metadata.get("free_subscription_id", "")
    if not subscription_id:
        raise RuntimeError("No Razorpay subscription ID is available. Start the ₹299/mo checkout first.")
    data = get_razorpay_subscription(subscription_id)
    status = str(data.get("status", "")).lower()
    if status == "active":
        update_free_billing_status("active", subscription_id)
        return True, status
    update_free_billing_status(status, subscription_id)
    return False, status


def friendly_auth_error(error) -> str:
    """Convert Supabase auth errors into user-friendly messages."""
    message = str(getattr(error, "message", error))
    lowered = message.lower()

    if "invalid login credentials" in lowered:
        return "Invalid email or password."
    if "email not confirmed" in lowered:
        return "Please verify your email address before signing in."
    if "user already registered" in lowered:
        return "An account with this email already exists. Please sign in."
    if "password should be at least" in lowered:
        return "Password must meet Supabase's minimum password requirements."
    if "rate limit" in lowered:
        return "Too many attempts. Please wait a moment and try again."
    if "name or service not known" in lowered:
        return (
            "Could not connect to Supabase. Check SUPABASE_URL and "
            "SUPABASE_ANON_KEY in Streamlit Secrets."
        )

    return message


def sign_up_user(full_name, company_name, email, password):
    """Create a persistent customer account in Supabase Auth."""
    supabase = get_supabase_client()

    return supabase.auth.sign_up(
        {
            "email": email.strip().lower(),
            "password": password,
            "options": {
                "data": {
                    "full_name": full_name.strip(),
                    "company_name": company_name.strip(),
                    "plan": "Free",
                }
            },
        }
    )


def sign_in_user(email, password):
    """Authenticate a customer using Supabase Auth."""
    supabase = get_supabase_client()

    return supabase.auth.sign_in_with_password(
        {
            "email": email.strip().lower(),
            "password": password,
        }
    )


def sign_out_user():
    """Sign the current user out of Supabase."""
    try:
        supabase = get_supabase_client()
        supabase.auth.sign_out()
    except Exception:
        pass


def set_authenticated_user(response):
    """Copy authenticated Supabase user information into session state."""
    user = getattr(response, "user", None)

    if user is None:
        raise RuntimeError(
            "Authentication succeeded but no user was returned."
        )

    metadata = getattr(user, "user_metadata", {}) or {}

    st.session_state.authenticated = True
    st.session_state.user_email = (user.email or "").lower()
    st.session_state.user_id = user.id
    st.session_state.user_name = metadata.get("full_name", "")
    st.session_state.company_name = metadata.get("company_name", "")
    st.session_state.user_plan = metadata.get("plan", "Free") or "Free"
    st.session_state.razorpay_subscription_id = metadata.get("razorpay_subscription_id", "")
    st.session_state.free_billing_status = metadata.get("free_billing_status", "")
    st.session_state.free_subscription_id = metadata.get("free_subscription_id", "")
    st.session_state.industry = metadata.get("industry", "BPO") or "BPO"

    # Supabase sets this automatically when the account is created — used
    # to work out how many days are left in the Free trial.
    created_at = getattr(user, "created_at", None)
    st.session_state.account_created_at = str(created_at) if created_at else ""


def clear_authentication():
    sign_out_user()

    st.session_state.authenticated = False
    st.session_state.user_email = ""
    st.session_state.user_id = ""
    st.session_state.user_name = ""
    st.session_state.company_name = ""
    st.session_state.user_plan = "Free"
    st.session_state.show_plans = False
    st.session_state.razorpay_checkout_url = ""
    st.session_state.razorpay_subscription_id = ""
    st.session_state.account_created_at = ""
    st.session_state.free_billing_status = ""
    st.session_state.free_subscription_id = ""
    st.session_state.razorpay_checkout_url_free = ""
    st.session_state.industry = "BPO"
    st.session_state.manufacturing_file_name = ""
    st.session_state.manufacturing_result = None
    st.session_state.manufacturing_report_bytes = None
    st.session_state.manufacturing_report_generated_at = None

    clear_analysis()


def trial_days_remaining():
    """Days left in the Free trial, or None if unknown (e.g. not signed in)."""
    raw = st.session_state.get("account_created_at", "")
    if not raw:
        return None
    try:
        created = datetime.fromisoformat(str(raw).replace("Z", "+00:00"))
    except ValueError:
        return None
    now = datetime.now(created.tzinfo) if created.tzinfo else datetime.now()
    elapsed_days = (now - created).days
    return FREE_TRIAL_DAYS - elapsed_days


def get_plan_config(plan):
    configs = {
        "Free": {
            "max_mb": 5,
            "copilot": True,
            "pdf": True,
            "email": False,
            "automation": False,
            "price": "₹299/mo (15 days free)",
        },
        "Professional": {
            "max_mb": 25,
            "copilot": True,
            "pdf": True,
            "email": True,
            "automation": True,
            "price": "₹1,999/mo",
        },
        "Business": {
            "max_mb": 100,
            "copilot": True,
            "pdf": True,
            "email": True,
            "automation": True,
            "price": "Custom",
        },
    }

    return configs.get(plan, configs["Free"])


def clear_analysis():
    st.session_state.n8n_sent = False
    st.session_state.n8n_result = None
    st.session_state.copilot_answer = None
    st.session_state.last_question = ""
    st.session_state.file_name = ""
    st.session_state.analysis_result = None
    st.session_state.analysis_df = None
    st.session_state.report_pdf = None
    st.session_state.report_generated_at = None


def normalize_n8n_response(response):
    try:
        data = response.json()
    except ValueError:
        return {"answer": response.text}

    if isinstance(data, list) and data:
        data = data[0]

    return data if isinstance(data, dict) else {"answer": data}


def parse_ai_answer(data):
    answer = data

    if isinstance(data, dict):
        answer = (
            data.get("answer")
            or data.get("response")
            or data.get("output")
            or data.get("text")
            or data.get("message")
        )

    if isinstance(answer, str):
        text = answer.strip()

        if text.startswith("```"):
            text = text[3:].strip()
            if text.lower().startswith("json"):
                text = text[4:].strip()
            if text.endswith("```"):
                text = text[:-3].strip()

        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return text

    return answer


def build_data_template_bytes():
    """Build the operational-data upload template as an in-memory .xlsx file."""

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    NAVY = "071A3D"
    BLUE = "0757B8"
    LIGHT_BLUE = "EAF2FD"
    SAMPLE_FILL = "FFF6E5"
    GREY = "667085"

    ws = wb.active
    ws.title = "Operational_Data"

    headers = [
        ("Date", "Date", 14),
        ("Employee_ID", "Text", 14),
        ("Employee_Name", "Text", 18),
        ("Team", "Text", 14),
        ("Target", "Number", 10),
        ("Production", "Number", 12),
        ("AHT_Actual", "Number", 12),
        ("AHT_Target", "Number", 12),
        ("Quality_%", "Number (0-100)", 12),
        ("SLA_%", "Number (0-100)", 10),
        ("Attendance", "Text (Present/Absent)", 16),
        ("Error_Count", "Number", 12),
        ("Error_Category", "Text", 18),
    ]

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="D9DEE7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (name, _, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    hint_font = Font(name="Arial", italic=True, color=GREY, size=9)
    hint_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for col_idx, (name, hint, _) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = hint_font
        cell.fill = hint_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    base_date = datetime(2026, 8, 3).date()
    sample_rows = [
        [base_date, "EMP-1001", "Aditi Sharma", "Collections", 50, 47, 6.4, 6.0, 96.5, 98.1, "Present", 1, "Documentation"],
        [base_date, "EMP-1002", "Rahul Verma", "Collections", 50, 41, 7.8, 6.0, 91.2, 93.4, "Present", 4, "Process error"],
        [base_date, "EMP-1003", "Meera Iyer", "Customer Care", 45, 46, 5.9, 6.0, 98.0, 99.0, "Absent", 0, ""],
        [base_date + timedelta(days=1), "EMP-1001", "Aditi Sharma", "Collections", 50, 44, 6.7, 6.0, 95.0, 97.2, "Present", 2, "Documentation"],
        [base_date + timedelta(days=1), "EMP-1004", "Karan Malhotra", "Customer Care", 45, 38, 8.2, 6.0, 88.5, 90.0, "Present", 6, "Escalation delay"],
    ]

    sample_font = Font(name="Arial", size=10)
    for r_offset, row_data in enumerate(sample_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=r_offset, column=col_idx, value=value)
            cell.font = sample_font
            cell.fill = PatternFill("solid", fgColor=SAMPLE_FILL)
            cell.border = border
            if headers[col_idx - 1][0] == "Date":
                cell.number_format = "DD-MMM-YYYY"

    ws.freeze_panes = "A3"

    note_row = len(sample_rows) + 4
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            "↑ Rows 3–7 are SAMPLE data showing the expected format. "
            "Delete them and paste your own operational data starting at row 3."
        ),
    )
    note.font = Font(name="Arial", italic=True, size=9, color=GREY)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))

    ins = wb.create_sheet("Instructions")
    ins.column_dimensions["A"].width = 20
    ins.column_dimensions["B"].width = 70
    ins.column_dimensions["C"].width = 14

    title = ins.cell(row=1, column=1, value="AI Operations Manager — Data Upload Template")
    title.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    ins.merge_cells("A1:C1")

    sub = ins.cell(
        row=2,
        column=1,
        value="Fill in the 'Operational_Data' sheet with your own rows, then upload the file (.xlsx or .csv).",
    )
    sub.font = Font(name="Arial", italic=True, size=10, color=GREY)
    ins.merge_cells("A2:C2")

    col_head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    col_head_fill = PatternFill("solid", fgColor=BLUE)
    for i, h in enumerate(["Column", "What to enter", "Required?"], start=1):
        c = ins.cell(row=4, column=i, value=h)
        c.font = col_head_font
        c.fill = col_head_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    field_docs = [
        ("Date", "The date the record applies to (one row per employee per day).", "Optional"),
        ("Employee_ID", "A unique ID for the employee (e.g. EMP-1001). Used to track the same person across days.", "Required"),
        ("Employee_Name", "The employee's full name, as it should appear in reports.", "Required"),
        ("Team", "The team or department the employee belongs to (e.g. Collections, Customer Care).", "Required"),
        ("Target", "The expected production/output target for that day (a number).", "Required"),
        ("Production", "The actual production/output achieved that day (a number).", "Required"),
        ("AHT_Actual", "Actual Average Handling Time for that day (in minutes, or your standard unit).", "Required"),
        ("AHT_Target", "The target Average Handling Time to compare against.", "Optional"),
        ("Quality_%", "Quality score for the day, as a percentage (0–100, not a decimal fraction).", "Required"),
        ("SLA_%", "SLA adherence for the day, as a percentage (0–100).", "Required"),
        ("Attendance", "The day's attendance status as text, e.g. 'Present' or 'Absent' (the app calculates absence % from this — don't enter a percentage here).", "Required"),
        ("Error_Count", "Number of errors recorded that day.", "Required"),
        ("Error_Category", "A short label for the main error type, if any (e.g. Documentation, Process error).", "Optional"),
    ]

    row_font = Font(name="Arial", size=10)
    req_font = Font(name="Arial", size=10, bold=True, color="B42318")
    opt_font = Font(name="Arial", size=10, color=GREY)

    r = 5
    for name, desc, required in field_docs:
        ins.cell(row=r, column=1, value=name).font = Font(name="Arial", bold=True, size=10, color=NAVY)
        ins.cell(row=r, column=2, value=desc).font = row_font
        ins.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        req_cell = ins.cell(row=r, column=3, value=required)
        req_cell.font = req_font if required == "Required" else opt_font
        ins.row_dimensions[r].height = 30
        r += 1

    r += 1
    notes_title = ins.cell(row=r, column=1, value="Notes")
    notes_title.font = Font(name="Arial", bold=True, size=12, color=NAVY)
    r += 1
    notes = [
        "• Keep the header row (row 1) exactly as provided — column names must match for the upload to work.",
        "• One row = one employee's record for one day. Add as many rows as you need.",
        "• Quality_% and SLA_% should be plain numbers like 96.5, not '96.5%' as text.",
        "• Attendance is a text status per row (e.g. 'Present' or 'Absent') — the app calculates absence % from this automatically.",
        "• Keep the sheet named 'Operational_Data' if using an Excel file with multiple sheets.",
        "• Free plan supports files up to 5 MB; Professional up to 25 MB; Business up to 100 MB.",
    ]
    for note_line in notes:
        ins.cell(row=r, column=1, value=note_line).font = Font(name="Arial", size=10)
        ins.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_procurement_report_bytes(prs, result):
    """
    Build the Manufacturing procurement comparison report (Overview, PR
    Summary, Item Detail sheets) as in-memory .xlsx bytes for download.
    Mirrors the standalone report generator used to validate this engine.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    overall = result["overall"]
    pr_rows = result["pr_rows"]
    item_rows = result["item_rows"]

    NAVY = "071A3D"
    LIGHT_BLUE = "EAF2FD"
    RISK_FILL = "FDEBEC"
    RISK_FONT = "B42318"
    GOOD_FILL = "EAF7EE"
    GREY = "667085"

    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D9DEE7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=NAVY)
    body_font = Font(name="Arial", size=10)

    # --- Overview ---
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22

    title = ws.cell(row=1, column=1, value="Procurement Price Comparison — Overview")
    title.font = Font(name="Arial", bold=True, size=14, color=NAVY)
    ws.merge_cells("A1:B1")

    sub = ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} · {overall['total_prs']} Purchase Requisitions analyzed")
    sub.font = Font(name="Arial", italic=True, size=9, color=GREY)
    ws.merge_cells("A2:B2")

    metrics = [
        ("Purchase Requisitions", overall["total_prs"]),
        ("Total Items", overall["total_items"]),
        ("Recommended Spend (₹)", overall["total_recommended_spend"]),
        ("Highest-Quote Spend (₹)", overall["total_highest_quote_spend"]),
        ("Potential Savings (₹)", overall["total_potential_savings"]),
        ("Overall Savings %", overall["overall_savings_pct"]),
        ("Single-Vendor Items (Risk)", overall["total_single_vendor_items"]),
        ("Price-Increase Items (Risk)", overall["total_price_increase_items"]),
        ("No-Quote Items (Risk)", overall["total_no_quote_items"]),
    ]
    r = 4
    for label, value in metrics:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, size=10, color=NAVY)
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = body_font
        if "Spend" in label or "Savings (" in label:
            vc.number_format = "₹#,##0"
        elif "%" in label:
            vc.number_format = "0.0"
        for cell in (lc, vc):
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE) if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        r += 1

    note = ws.cell(row=r + 1, column=1, value=(
        "Recommended Vendor = lowest quoted total price per item. Savings = difference vs the "
        "highest quote received for that item. Risk flags call out items with only one quote, "
        "a >10% price increase vs the last recorded order, or no valid quote at all."
    ))
    note.font = Font(name="Arial", italic=True, size=9, color=GREY)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=2)
    ws.row_dimensions[r + 1].height = 45

    # Executive summary — pulled from procurement_engine.build_insights so
    # the report always matches exactly what was shown in-app.
    _insights = procurement_engine.build_insights(result)
    _risk_level = _insights["risk_level"].split(" ", 1)[-1]  # drop the emoji for the plain-text report title

    r_summary = r + 3
    summary_title = ws.cell(row=r_summary, column=1, value=f"Executive Summary — {_risk_level}")
    summary_title.font = Font(name="Arial", bold=True, size=12, color=NAVY)
    ws.merge_cells(start_row=r_summary, start_column=1, end_row=r_summary, end_column=2)
    r_summary += 1

    for point in _insights["summary_points"]:
        # Strip the leading emoji for a plain-text report line.
        line = point.split(" ", 1)[-1] if point[:1] in "🟢🟡🟠🔴⚪" else point
        cell = ws.cell(row=r_summary, column=1, value=f"• {line}")
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r_summary, start_column=1, end_row=r_summary, end_column=2)
        ws.row_dimensions[r_summary].height = 28
        r_summary += 1

    r_summary += 1
    rec_title = ws.cell(row=r_summary, column=1, value="Recommendation")
    rec_title.font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ws.merge_cells(start_row=r_summary, start_column=1, end_row=r_summary, end_column=2)
    r_summary += 1
    rec_cell = ws.cell(row=r_summary, column=1, value=_insights["recommendation"])
    rec_cell.font = Font(name="Arial", size=10)
    rec_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r_summary, start_column=1, end_row=r_summary, end_column=2)
    ws.row_dimensions[r_summary].height = 45

    # --- PR Summary ---
    ws2 = wb.create_sheet("PR Summary")
    pr_headers = ["PR Sheet", "Plant / Indenter", "Working Date", "Vendors Compared", "Items",
                  "Recommended Spend", "Highest-Quote Spend", "Potential Savings", "Savings %",
                  "Single-Vendor Items", "Price-Increase Items", "No-Quote Items"]
    for c, h in enumerate(pr_headers, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws2.row_dimensions[1].height = 30
    for i, w in enumerate([12, 46, 14, 16, 8, 16, 18, 16, 10, 16, 16, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    for ridx, row in enumerate(pr_rows, start=2):
        values = [
            row["PR Sheet"], row["Plant / Indenter"], row["Working Date"], row["Vendors Compared"],
            row["Items"], row["Recommended Spend"], row["Highest-Quote Spend"], row["Potential Savings"],
            row["Savings %"], row["Single-Vendor Items (Risk)"], row["Price-Increase Items (Risk)"],
            row["No-Quote Items (Risk)"],
        ]
        for c, v in enumerate(values, start=1):
            cell = ws2.cell(row=ridx, column=c, value=v)
            cell.font = body_font
            cell.border = border
            if c in (6, 7, 8):
                cell.number_format = "₹#,##0"
            if c == 9 and v is not None:
                cell.number_format = "0.0"
            if c in (10, 11, 12) and v:
                cell.fill = PatternFill("solid", fgColor=RISK_FILL)
                cell.font = Font(name="Arial", size=10, color=RISK_FONT, bold=True)
    ws2.freeze_panes = "A2"

    # --- Item Detail ---
    ws3 = wb.create_sheet("Item Detail")
    item_headers = ["PR Sheet", "PR No", "Item Code", "Item Name", "Qty", "UM", "Vendors Quoted",
                     "Recommended Vendor", "Recommended Total Price", "2nd Best Vendor", "2nd Best Total Price",
                     "Highest Quoted Vendor", "Highest Total Price", "Savings vs Highest", "Savings vs Highest %",
                     "Previous Unit Price", "Previous Vendor", "% Change vs Previous", "Risk Flags"]
    for c, h in enumerate(item_headers, start=1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws3.row_dimensions[1].height = 32
    for i, w in enumerate([10, 12, 12, 42, 9, 7, 10, 24, 16, 22, 15, 24, 15, 14, 12, 14, 20, 14, 32], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    money_cols = {9, 11, 13, 14, 16}
    pct_cols = {15, 18}
    for ridx, row in enumerate(item_rows, start=2):
        values = [
            row["PR Sheet"], row["PR No"], row["Item Code"], row["Item Name"], row["Qty"], row["UM"],
            row["Vendors Quoted"], row["Recommended Vendor"], row["Recommended Total Price"],
            row["2nd Best Vendor"], row["2nd Best Total Price"], row["Highest Quoted Vendor"],
            row["Highest Total Price"], row["Savings vs Highest"], row["Savings vs Highest %"],
            row["Previous Unit Price"], row["Previous Vendor"], row["% Change vs Previous"], row["Risk Flags"],
        ]
        has_risk = bool(row["Risk Flags"])
        for c, v in enumerate(values, start=1):
            cell = ws3.cell(row=ridx, column=c, value=v)
            cell.font = body_font
            cell.border = border
            if c in money_cols and v is not None:
                cell.number_format = "₹#,##0.00"
            if c in pct_cols and v is not None:
                cell.number_format = "0.0"
            if c == 19 and has_risk:
                cell.fill = PatternFill("solid", fgColor=RISK_FILL)
                cell.font = Font(name="Arial", size=10, color=RISK_FONT, bold=True)
            elif c == 8 and not has_risk and row["Vendors Quoted"] > 1:
                cell.fill = PatternFill("solid", fgColor=GOOD_FILL)
    ws3.freeze_panes = "B2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def render_footer():
    st.divider()
    st.markdown(
        f"""<div class="gi-footer">
<strong>Generative Insight</strong> · AI Operations Copilot v{APP_VERSION}
<br>
Insights today. Intelligence tomorrow.
<br>
<a class="website-link" href="{WEBSITE_URL}" target="_blank" rel="noopener noreferrer">generativeinsight.in</a>
&nbsp;·&nbsp;
© {datetime.now().year}
</div>""",
        unsafe_allow_html=True,
    )


def render_coming_soon_flow(industry_key):
    """
    Shown for any industry selected in the dropdown that doesn't have a
    real engine module yet (Retail, Logistics, Healthcare, ...). Never
    shows fake data or placeholder insights — just a clear message and
    a way to request the module, same brand styling as the rest of the app.
    """
    show_brand_header(compact=True)

    label = INDUSTRY_LABELS.get(industry_key, industry_key)

    st.markdown(
        f'<div class="main-title">{label}</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""<div class="hero">
<h3>This industry module is coming soon</h3>
<p class="small-muted">{label} isn't built yet — Manufacturing (Procurement) and BPO / Call Center are live today. Switch back to one of those from the sidebar, or tell us what this module needs.</p>
</div>""",
        unsafe_allow_html=True,
    )

    st.markdown("### What we'd need to build this")
    st.write(
        "Every industry module follows the same pattern as the existing "
        "ones: a sample template of your operational data, the KPIs that "
        "matter for this industry, and how risk should be flagged. Share "
        "that and we can turn it into a working module the same way "
        "Manufacturing (Procurement) was built."
    )

    contact_email = secret("SUPPORT_EMAIL", "hello@generativeinsight.in")
    st.link_button(
        f"📧 Request the {label} module",
        f"mailto:{contact_email}?subject=Requesting%20the%20{label.replace(' ', '%20')}%20module",
        use_container_width=False,
    )

    render_footer()


def render_manufacturing_flow(plan_config):
    """
    Manufacturing industry flow: upload a Price Comparative Sheet workbook
    (one sheet per Purchase Requisition, multiple vendor quote blocks per
    sheet), get vendor-comparison analysis, recommended vendor per item,
    savings vs alternatives, and risk flags. Uses procurement_engine.py,
    a self-contained module with no dependency on engine.py / analyze_data.
    """

    st.subheader("🏭 Procurement Setup")

    col1, col2 = st.columns(2)
    with col1:
        company_name = st.text_input(
            "Company Name",
            value=st.session_state.get("company_name", ""),
            placeholder="e.g. ABC Manufacturing Pvt. Ltd.",
            key="mfg_company_name",
        )
    with col2:
        report_name = st.text_input(
            "Report Name",
            value="Procurement Price Comparison",
            key="mfg_report_name",
        )

    uploaded = st.file_uploader(
        f"📁 Upload Price Comparative Sheet workbook (.xlsx) — max {plan_config['max_mb']} MB",
        type=["xlsx"],
        key="mfg_file_uploader",
    )

    if not uploaded:
        st.info(
            "Upload a Price Comparative Sheet workbook to compare vendor quotes, "
            "identify the best price per item, and flag procurement risks."
        )
        st.markdown("### Expected format")
        st.caption(
            "One sheet per Purchase Requisition. Each sheet needs a header row with "
            "S.N., PR NO, Item Code, Item Name, Qty, UM, then one 5-column block per "
            "vendor (Make, Quoted Price, Dis %, Dis Price, Total Price), followed by "
            "Doc. Date, Vendor, and DPTPL (previous order unit price) columns."
        )
        st.markdown("### What you get")
        a, b, c, d = st.columns(4)
        a.metric("Vendor Comparison", "✓")
        b.metric("Recommended Vendor", "✓")
        c.metric("Savings Analysis", "✓")
        d.metric("Risk Flags", "✓")
        render_footer()
        return

    file_mb = uploaded.size / (1024 * 1024)
    if file_mb > plan_config["max_mb"]:
        st.error(
            f"File is {file_mb:.2f} MB. Your {st.session_state.user_plan} plan "
            f"supports files up to {plan_config['max_mb']} MB."
        )
        render_footer()
        return

    if st.session_state.manufacturing_file_name != uploaded.name:
        st.session_state.manufacturing_file_name = uploaded.name
        st.session_state.manufacturing_result = None
        st.session_state.manufacturing_report_bytes = None

    _mfg_start_time = time.time()

    try:
        with st.spinner("🔄 Fetching and reading your procurement workbook..."):
            prs = procurement_engine.parse_workbook(uploaded)
            if not prs:
                st.error(
                    "❌ No Price Comparative Sheet-style data found in this workbook. "
                    "Check that at least one sheet has a header row starting with 'S.N.'"
                )
                render_footer()
                return
        with st.spinner("🧠 Comparing vendor quotes and flagging risk..."):
            result = procurement_engine.analyze_procurement(prs)
    except Exception as e:
        st.error(f"❌ Could not analyze the uploaded file: {e}")
        render_footer()
        return

    _mfg_elapsed = time.time() - _mfg_start_time

    st.session_state.manufacturing_result = result
    overall = result["overall"]

    st.markdown(
        f"""<div class="hero">
<h3>Procurement Overview</h3>
<p class="small-muted">{company_name or "Your organization"} · {report_name}</p>
<p class="small-muted">⏱️ Data fetched and analyzed in {_mfg_elapsed:.2f}s · {overall['total_items']} item(s) across {overall['total_prs']} PR sheet(s)</p>
</div>""",
        unsafe_allow_html=True,
    )

    r1, r2, r3, r4 = st.columns(4)
    with r1:
        render_metric_card("💰", "Recommended Spend", f"₹{overall['total_recommended_spend']:,.0f}")
    with r2:
        render_metric_card("🔎", "Highest-Quote Spend", f"₹{overall['total_highest_quote_spend']:,.0f}")
    with r3:
        render_metric_card(
            "📉", "Potential Savings", f"₹{overall['total_potential_savings']:,.0f}",
            delta=f"{overall['overall_savings_pct']}%" if overall['overall_savings_pct'] is not None else None,
            delta_tone="good" if (overall['overall_savings_pct'] or 0) > 0 else "neutral",
        )
    with r4:
        render_metric_card("📋", "Purchase Requisitions", overall["total_prs"])

    risk1, risk2, risk3 = st.columns(3)
    with risk1:
        render_metric_card(
            "⚠️", "Single-Vendor Items", overall["total_single_vendor_items"],
            delta_tone="bad" if overall["total_single_vendor_items"] > 0 else "good",
        )
    with risk2:
        render_metric_card(
            "📈", "Price-Increase Items", overall["total_price_increase_items"],
            delta_tone="bad" if overall["total_price_increase_items"] > 0 else "good",
        )
    with risk3:
        render_metric_card(
            "❓", "No-Quote Items", overall["total_no_quote_items"],
            delta_tone="bad" if overall["total_no_quote_items"] > 0 else "good",
        )

    # ============================================================
    # EXECUTIVE SUMMARY — risk level, summary bullets, and the
    # recommendation are all computed once in procurement_engine.py
    # (single source of truth shared with the Excel report and the
    # AI Copilot context), mirroring the BPO engine's insight pattern.
    # ============================================================

    mfg_insights = procurement_engine.build_insights(result)
    mfg_risk_level = mfg_insights["risk_level"]
    mfg_summary_points = mfg_insights["summary_points"]
    mfg_recommendation = mfg_insights["recommendation"]

    st.subheader("🧠 Executive Summary")
    st.metric("Procurement Risk", mfg_risk_level)
    for point in mfg_summary_points:
        st.write(point)
    st.info(f"💡 **Procurement Recommendation:** {mfg_recommendation}")

    mfg_tabs = st.tabs(["📋 PR Summary", "📦 Item Detail", "⚠️ Risk Items", "🤖 Procurement Copilot", "📄 Report", "💳 Billing"])

    with mfg_tabs[0]:
        st.subheader("Purchase Requisition Summary")
        pr_df = pd.DataFrame(result["pr_rows"])
        if not pr_df.empty:
            st.dataframe(pr_df, use_container_width=True, hide_index=True)
        else:
            st.info("No PR-level data available.")

    with mfg_tabs[1]:
        st.subheader("Item-Level Comparison")
        item_df = pd.DataFrame(result["item_rows"])
        if not item_df.empty:
            st.dataframe(item_df, use_container_width=True, hide_index=True)
        else:
            st.info("No item-level data available.")

    with mfg_tabs[2]:
        st.subheader("⚠️ Flagged Items")
        item_df = pd.DataFrame(result["item_rows"])
        if not item_df.empty:
            risk_df = item_df[item_df["Risk Flags"].astype(str).str.len() > 0]
            if not risk_df.empty:
                st.dataframe(risk_df, use_container_width=True, hide_index=True)
            else:
                st.success("✅ No flagged items — every item has multiple quotes with no unusual price movement.")
        else:
            st.info("No item-level data available.")

    with mfg_tabs[3]:
        st.subheader("🤖 Procurement Copilot")
        st.caption(
            "Ask questions about this procurement data. The Copilot is "
            "instructed to use only the supplied vendor comparison context."
        )

        mfg_question = st.text_input(
            "Ask your procurement question",
            placeholder="Which items have the biggest savings opportunity and which vendor should we use?",
            key="mfg_copilot_question",
        )

        mfg_ask_copilot = st.button(
            "🚀 Ask Procurement Copilot",
            type="primary",
            use_container_width=True,
            key="mfg_ask_copilot",
        )

        mfg_copilot_url = secret("N8N_COPILOT_WEBHOOK_URL")

        if mfg_ask_copilot:
            if not mfg_question.strip():
                st.warning("⚠️ Please enter a question first.")
            elif not mfg_copilot_url:
                st.error("❌ N8N_COPILOT_WEBHOOK_URL is not configured in Streamlit Secrets.")
            else:
                mfg_context = (
                    f"Company:\n{company_name}\n\nReport:\n{report_name}\n\n"
                    + procurement_engine.make_ai_prompt(result, mfg_insights)
                )
                mfg_payload = {
                    "question": mfg_question.strip(),
                    "company_name": company_name.strip(),
                    "report_name": report_name.strip(),
                    "context": mfg_context,
                    "user_id": st.session_state.user_id,
                    "user_email": st.session_state.user_email,
                }
                try:
                    with st.spinner("🤖 Procurement Copilot is analyzing..."):
                        mfg_copilot_response = requests.post(
                            mfg_copilot_url,
                            json=mfg_payload,
                            headers={"Content-Type": "application/json"},
                            timeout=120,
                        )
                    if mfg_copilot_response.status_code < 300:
                        mfg_raw = normalize_n8n_response(mfg_copilot_response)
                        st.session_state.copilot_answer = parse_ai_answer(mfg_raw)
                        st.session_state.last_question = mfg_question.strip()
                    else:
                        st.session_state.copilot_answer = None
                        st.error(f"❌ Copilot workflow failed: HTTP {mfg_copilot_response.status_code}")
                        st.code(mfg_copilot_response.text, language="text")
                except requests.exceptions.Timeout:
                    st.session_state.copilot_answer = None
                    st.error("⏱️ Procurement Copilot timed out. Please try again.")
                except requests.exceptions.RequestException as e:
                    st.session_state.copilot_answer = None
                    st.error(f"❌ Copilot request failed: {e}")

        if st.session_state.copilot_answer:
            st.divider()
            st.markdown("### 🧠 Copilot Analysis")
            st.caption("Question: " + st.session_state.last_question)
            mfg_answer = st.session_state.copilot_answer
            if isinstance(mfg_answer, dict):
                if mfg_answer.get("what_is_happening"):
                    st.markdown("#### 🔎 What is happening")
                    st.info(mfg_answer["what_is_happening"])
                if mfg_answer.get("recommended_actions"):
                    st.markdown("#### ✅ Recommended Actions")
                    for i, action in enumerate(mfg_answer["recommended_actions"], 1):
                        st.markdown(f"**{i}.** {action}")
            elif isinstance(mfg_answer, str):
                st.markdown(mfg_answer)
            else:
                st.code(str(mfg_answer), language="text")

    with mfg_tabs[4]:
        st.subheader("📄 Procurement Report")
        if st.button("📄 Generate Report", type="primary", use_container_width=True, key="mfg_generate_report"):
            try:
                with st.spinner("Generating procurement report..."):
                    report_bytes = build_procurement_report_bytes(prs, result)
                st.session_state.manufacturing_report_bytes = report_bytes
                st.session_state.manufacturing_report_generated_at = datetime.now()
            except Exception as e:
                st.error(f"❌ Could not generate report: {e}")

        if st.session_state.manufacturing_report_bytes:
            filename = (
                f"{company_name or 'procurement'}_comparison_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            )
            st.download_button(
                "⬇️ Download Procurement Report (.xlsx)",
                data=st.session_state.manufacturing_report_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="mfg_download_report",
            )
            if st.session_state.manufacturing_report_generated_at:
                st.caption(
                    "Generated "
                    + st.session_state.manufacturing_report_generated_at.strftime("%d %b %Y, %H:%M")
                )

    with mfg_tabs[5]:
        st.subheader("💳 Subscription & Billing")
        st.info(f"You are currently using the **{st.session_state.user_plan}** plan.")
        show_pricing("billing_manufacturing")

    with st.expander("🧠 AI Analyst Context / Prompt", expanded=False):
        st.code(procurement_engine.make_ai_prompt(result, mfg_insights), language="text")

    render_footer()


def dataframe_to_text(df):
    if df is None or df.empty:
        return "No records available."

    return df.to_string(index=False)


def build_copilot_context(
    company_name,
    report_name,
    result,
    productivity_target,
    quality_target,
    sla_target,
    aht_target,
    risk_level,
    summary_points,
):
    overall = result["overall"]

    return f"""
Company:
{company_name}

Report:
{report_name}

Operational KPIs:

Productivity:
{float(overall["productivity"]):.2f}% | Target: {productivity_target}%

Quality:
{float(overall["quality"]):.2f}% | Target: {quality_target}%

SLA:
{float(overall["sla"]):.2f}% | Target: {sla_target}%

Average AHT:
{float(overall["aht"]):.2f} | Target: {aht_target}

Overall Risk:
{risk_level}

Operational Findings:
{dataframe_to_text(result.get("findings"))}

Recommended Actions:
{dataframe_to_text(result.get("actions"))}

Employee Risk Data:
{dataframe_to_text(result.get("employees"))}

Team Performance:
{dataframe_to_text(result.get("team"))}

KPI Summary:
{chr(10).join(summary_points)}
""".strip()


def create_pdf_report(
    company_name,
    report_name,
    result,
    risk_level,
    summary_points,
    recommendation,
):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import (
            getSampleStyleSheet,
            ParagraphStyle,
        )
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        raise RuntimeError(
            "PDF generation requires reportlab. "
            "Add reportlab to requirements.txt."
        )

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"{company_name} - {report_name}",
        author=APP_NAME,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "GI_Title",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        spaceAfter=8,
    )

    heading_style = ParagraphStyle(
        "GI_Heading",
        parent=styles["Heading2"],
        fontSize=13,
        leading=16,
        spaceBefore=10,
        spaceAfter=6,
    )

    body_style = ParagraphStyle(
        "GI_Body",
        parent=styles["BodyText"],
        fontSize=9,
        leading=12,
    )

    story = []

    story.append(Paragraph("Generative Insight", title_style))
    story.append(
        Paragraph(
            f"<b>{company_name}</b> — {report_name}<br/>"
            f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}",
            body_style,
        )
    )

    story.append(Spacer(1, 8))

    overall = result["overall"]

    kpi_rows = [
        ["KPI", "Actual", "Target"],
        [
            "Productivity",
            f'{float(overall["productivity"]):.2f}%',
            f'{result.get("_targets", {}).get("productivity", "")}%',
        ],
        [
            "Quality",
            f'{float(overall["quality"]):.2f}%',
            f'{result.get("_targets", {}).get("quality", "")}%',
        ],
        [
            "SLA",
            f'{float(overall["sla"]):.2f}%',
            f'{result.get("_targets", {}).get("sla", "")}%',
        ],
        [
            "Average AHT",
            f'{float(overall["aht"]):.2f}',
            f'{result.get("_targets", {}).get("aht", "")}',
        ],
    ]

    story.append(
        Paragraph("Executive Overview", heading_style)
    )

    story.append(
        Paragraph(
            f"<b>Risk:</b> {risk_level}",
            body_style,
        )
    )

    story.append(Spacer(1, 5))

    table = Table(
        kpi_rows,
        colWidths=[55 * mm, 45 * mm, 45 * mm],
    )

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)

    story.append(
        Paragraph("KPI Summary", heading_style)
    )

    for item in summary_points:
        cleaned = (
            item.replace("🔴 ", "")
            .replace("🟢 ", "")
            .replace("🟠 ", "")
            .replace("🟡 ", "")
        )
        story.append(
            Paragraph(cleaned, body_style)
        )

    story.append(
        Paragraph("Management Recommendation", heading_style)
    )

    story.append(
        Paragraph(
            recommendation,
            body_style,
        )
    )

    for title, key in [
        ("Team Performance", "team"),
        ("Operational Findings", "findings"),
        ("Recommended Actions", "actions"),
        ("Employee Risk", "employees"),
    ]:
        df = result.get(key)

        if isinstance(df, pd.DataFrame) and not df.empty:
            story.append(
                Paragraph(title, heading_style)
            )

            pdf_df = df.copy()

            if len(pdf_df.columns) > 8:
                pdf_df = pdf_df.iloc[:, :8]

            headers = [str(c) for c in pdf_df.columns]
            rows = [headers]

            for _, row in pdf_df.head(50).iterrows():
                rows.append(
                    [str(v)[:90] for v in row.tolist()]
                )

            col_count = len(headers)
            available_width = 180 * mm
            col_width = available_width / max(col_count, 1)

            tbl = Table(
                rows,
                colWidths=[col_width] * col_count,
                repeatRows=1,
            )

            tbl.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#111827"),
                        ),
                        (
                            "TEXTCOLOR",
                            (0, 0),
                            (-1, 0),
                            colors.white,
                        ),
                        (
                            "GRID",
                            (0, 0),
                            (-1, -1),
                            0.25,
                            colors.grey,
                        ),
                        (
                            "FONTSIZE",
                            (0, 0),
                            (-1, -1),
                            6,
                        ),
                        (
                            "VALIGN",
                            (0, 0),
                            (-1, -1),
                            "TOP",
                        ),
                    ]
                )
            )

            story.append(tbl)
            story.append(Spacer(1, 5))

    story.append(Spacer(1, 8))

    story.append(
        Paragraph(
            "Generated by Generative Insight AI Operations Copilot. "
            "AI recommendations should be validated against operational "
            "evidence before management action.",
            body_style,
        )
    )

    doc.build(story)

    buffer.seek(0)
    return buffer.getvalue()


def send_email_report(
    recipient,
    subject,
    body,
    pdf_bytes=None,
    pdf_filename="operations_report.pdf",
):
    smtp_host = secret("SMTP_HOST")
    smtp_port = int(secret("SMTP_PORT", "587"))
    smtp_user = secret("SMTP_USERNAME")
    smtp_password = secret("SMTP_PASSWORD")
    smtp_from = secret("SMTP_FROM", smtp_user)

    if not all(
        [smtp_host, smtp_user, smtp_password, smtp_from]
    ):
        raise RuntimeError(
            "SMTP settings are not configured in Streamlit Secrets."
        )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_from
    message["To"] = recipient

    message.set_content(body)

    if pdf_bytes:
        message.add_attachment(
            pdf_bytes,
            maintype="application",
            subtype="pdf",
            filename=pdf_filename,
        )

    with smtplib.SMTP(
        smtp_host,
        smtp_port,
        timeout=30,
    ) as server:
        server.starttls()
        server.login(
            smtp_user,
            smtp_password,
        )
        server.send_message(message)


# ============================================================
# RAZORPAY SUBSCRIPTIONS
# ============================================================

RAZORPAY_API_BASE = "https://api.razorpay.com/v1"


def razorpay_is_configured(plan_id_secret="RAZORPAY_PROFESSIONAL_PLAN_ID"):
    return all(
        [
            secret("RAZORPAY_KEY_ID"),
            secret("RAZORPAY_KEY_SECRET"),
            secret(plan_id_secret),
        ]
    )


def create_razorpay_subscription(
    customer_email="",
    customer_name="",
    plan_id_secret="RAZORPAY_PROFESSIONAL_PLAN_ID",
    plan_label="Professional",
    total_count_secret="RAZORPAY_PROFESSIONAL_TOTAL_COUNT",
):
    key_id = secret("RAZORPAY_KEY_ID")
    key_secret = secret("RAZORPAY_KEY_SECRET")
    plan_id = secret(plan_id_secret)

    if not key_id or not key_secret or not plan_id:
        raise RuntimeError(
            f"Razorpay is not configured. Add RAZORPAY_KEY_ID, "
            f"RAZORPAY_KEY_SECRET and {plan_id_secret} "
            "to Streamlit Secrets."
        )

    raw_total_count = secret(total_count_secret, "12")
    try:
        total_count = int(raw_total_count)
    except (TypeError, ValueError):
        total_count = 12

    if total_count < 1:
        total_count = 12

    payload = {
        "plan_id": plan_id,
        "total_count": total_count,
        "customer_notify": 1,
        "notes": {
            "application": APP_NAME,
            "plan": plan_label,
            "customer_email": str(customer_email or "")[:255],
            "customer_name": str(customer_name or "")[:255],
        },
    }

    try:
        response = requests.post(
            f"{RAZORPAY_API_BASE}/subscriptions",
            auth=(key_id, key_secret),
            json=payload,
            timeout=30,
        )
    except requests.exceptions.Timeout as exc:
        raise RuntimeError(
            "Razorpay request timed out. Please try again."
        ) from exc
    except requests.exceptions.RequestException as exc:
        raise RuntimeError(
            f"Could not connect to Razorpay: {exc}"
        ) from exc

    try:
        data = response.json()
    except ValueError:
        data = {"error": response.text}

    if response.status_code >= 300:
        error_message = data.get("error", data) if isinstance(data, dict) else data
        if isinstance(error_message, dict):
            error_message = (
                error_message.get("description")
                or error_message.get("reason")
                or str(error_message)
            )
        raise RuntimeError(
            f"Razorpay subscription creation failed (HTTP {response.status_code}): "
            f"{error_message}"
        )

    checkout_url = data.get("short_url")
    subscription_id = data.get("id")

    if not checkout_url or not subscription_id:
        raise RuntimeError(
            "Razorpay created the subscription but did not return a valid "
            "subscription checkout URL."
        )

    return {
        "id": subscription_id,
        "status": data.get("status", "created"),
        "short_url": checkout_url,
        "plan_id": data.get("plan_id", plan_id),
    }


# ============================================================
# PRICING
# ============================================================

def show_pricing(section_id="default"):
    """Display the existing pricing UI with Razorpay Professional checkout."""

    st.markdown("### 💳 Plans")

    c1, c2, c3 = st.columns(3)

    plans = [
        (
            c1,
            "Free",
            "₹299/mo",
            [
                "Free for first 15 days",
                "5 MB file limit",
                "Dashboard analytics",
                "AI Copilot",
                "PDF report",
            ],
            (
                "Current plan"
                if st.session_state.user_plan == "Free"
                else "Start Free Trial"
            ),
        ),
        (
            c2,
            "Professional",
            "₹1,999/mo",
            [
                "25 MB file limit",
                "AI Copilot",
                "PDF + email reports",
                "n8n automation",
            ],
            "Current plan"
            if st.session_state.user_plan == "Professional"
            else "Upgrade",
        ),
        (
            c3,
            "Business",
            "Custom",
            [
                "100 MB file limit",
                "Advanced automation",
                "Custom workflows",
                "Team deployment",
            ],
            "Contact Sales",
        ),
    ]

    for col, name, price, features, button in plans:
        with col:
            plan_visuals = {
                "Free": ("🌱", "15 days free, then ₹299/mo"),
                "Professional": ("🚀", "Advanced intelligence & automation"),
                "Business": ("🏢", "Scale AI operations across teams"),
            }
            icon, description = plan_visuals.get(name, ("✨", "Operational intelligence"))
            popular_badge = '<div class="plan-badge-popular">Most popular</div>' if name == "Professional" else ""

            st.markdown(
                f'''<div class="plan-card">{popular_badge}
<div style="font-size:2.6rem; line-height:1; margin-bottom:.7rem;">{icon}</div>
<div class="plan-name" style="font-size:1.15rem; font-weight:800;">{name}</div>
<div class="plan-price" style="font-size:1.65rem; font-weight:850; margin:.35rem 0;">{price}</div>
<div class="plan-description" style="font-size:.86rem; margin-bottom:.8rem;">{description}</div>
{''.join(f'<div class="plan-feature" style="margin:.35rem 0;">✓ {feature}</div>' for feature in features)}
</div>''',
                unsafe_allow_html=True,
            )

            # Professional uses the Razorpay subscription API.
            # Free/Business retain the existing checkout-link behavior.
            if name == "Professional":
                if st.session_state.user_plan == "Professional":
                    st.button(
                        "Current plan",
                        use_container_width=True,
                        disabled=True,
                        key=f"professional_current_{section_id}",
                    )
                elif not razorpay_is_configured():
                    st.button(
                        "Upgrade",
                        use_container_width=True,
                        disabled=True,
                        key=f"professional_disabled_{section_id}",
                    )
                    st.caption(
                        "Razorpay checkout is not configured yet."
                    )
                else:
                    # Keep the payment option visible in every environment.
                    # Secure activation still requires an authenticated Supabase user.
                    if st.button(
                        "💳 Upgrade to Professional",
                        type="primary",
                        use_container_width=True,
                        key=f"razorpay_upgrade_{section_id}",
                    ):
                        ready, message = razorpay_activation_ready()
                        if not ready:
                            st.error(f"❌ {message}")
                        else:
                            subscription = None
                            try:
                                with st.spinner("Creating secure Razorpay subscription..."):
                                    subscription = create_razorpay_subscription(
                                        customer_email=st.session_state.get("user_email", ""),
                                        customer_name=st.session_state.get("user_name", ""),
                                    )
                            except Exception as exc:
                                st.session_state.razorpay_checkout_url = ""
                                st.error(f"❌ Razorpay could not create the subscription: {exc}")

                            if subscription is not None:
                                try:
                                    # Save tracking first. Checkout URL is exposed only after
                                    # Supabase tracking succeeds, eliminating the previous
                                    # "checkout created but subscription tracking could not be saved" state.
                                    update_user_plan(
                                        "Free",
                                        subscription["id"],
                                        subscription.get("status", "created"),
                                    )
                                    st.session_state.razorpay_checkout_url = subscription["short_url"]
                                    st.session_state.razorpay_subscription_id = subscription["id"]
                                    st.success("Subscription created. Continue to secure Razorpay checkout.")
                                except Exception as exc:
                                    st.session_state.razorpay_checkout_url = ""
                                    st.error(
                                        f"❌ Razorpay subscription {subscription['id']} was created, "
                                        f"but saving it to your account (Supabase) failed: {exc}\n\n"
                                        "This usually means SUPABASE_SERVICE_ROLE_KEY is missing, wrong, "
                                        "or swapped with the anon key in Streamlit Secrets."
                                    )

                if st.session_state.get("razorpay_checkout_url"):
                    st.link_button(
                        "💳 Continue to Razorpay Checkout",
                        st.session_state.razorpay_checkout_url,
                        use_container_width=True,
                    )
                    subscription_id = st.session_state.get(
                        "razorpay_subscription_id", ""
                    )
                    if subscription_id:
                        st.caption(f"Subscription ID: {subscription_id}")
                        if st.button("🔄 Verify Professional Payment", use_container_width=True, key=f"verify_razorpay_{section_id}"):
                            try:
                                with st.spinner("Verifying your Razorpay subscription..."):
                                    active, status = verify_professional_subscription()
                                if active:
                                    st.success("✅ Payment verified. Professional plan activated.")
                                    st.session_state.razorpay_checkout_url = ""
                                    st.rerun()
                                else:
                                    st.info(f"Payment is not active yet. Razorpay status: {status or 'unknown'}. Complete checkout and try again.")
                            except Exception as exc:
                                st.error(f"❌ Verification failed: {exc}")

            elif name == "Free":
                if free_billing_is_active():
                    st.button(
                        "✅ Active (₹299/mo)",
                        use_container_width=True,
                        disabled=True,
                        key=f"free_active_{section_id}",
                    )
                elif st.session_state.user_plan != "Free":
                    st.button(
                        "Included",
                        use_container_width=True,
                        disabled=True,
                        key=f"free_included_{section_id}",
                    )
                elif not razorpay_is_configured("RAZORPAY_FREE_PLAN_ID"):
                    st.button(
                        button,
                        use_container_width=True,
                        disabled=True,
                        key=f"free_disabled_{section_id}",
                    )
                    st.caption("Razorpay checkout for this plan is not configured yet.")
                else:
                    _remaining = trial_days_remaining()
                    _label = (
                        "💳 Continue for ₹299/mo"
                        if _remaining is not None and _remaining <= 0
                        else "💳 Start Free Trial"
                    )
                    if st.button(
                        _label,
                        use_container_width=True,
                        key=f"razorpay_free_{section_id}",
                    ):
                        ready, message = razorpay_activation_ready("RAZORPAY_FREE_PLAN_ID", "Free")
                        if not ready:
                            st.error(f"❌ {message}")
                        else:
                            subscription = None
                            try:
                                with st.spinner("Creating secure Razorpay subscription..."):
                                    subscription = create_razorpay_subscription(
                                        customer_email=st.session_state.get("user_email", ""),
                                        customer_name=st.session_state.get("user_name", ""),
                                        plan_id_secret="RAZORPAY_FREE_PLAN_ID",
                                        plan_label="Free",
                                        total_count_secret="RAZORPAY_FREE_TOTAL_COUNT",
                                    )
                            except Exception as exc:
                                st.session_state.razorpay_checkout_url_free = ""
                                st.error(f"❌ Razorpay could not create the subscription: {exc}")

                            if subscription is not None:
                                try:
                                    update_free_billing_status(
                                        subscription.get("status", "created"),
                                        subscription["id"],
                                    )
                                    st.session_state.razorpay_checkout_url_free = subscription["short_url"]
                                    st.success("Subscription created. Continue to secure Razorpay checkout.")
                                except Exception as exc:
                                    st.session_state.razorpay_checkout_url_free = ""
                                    st.error(
                                        f"❌ Razorpay subscription {subscription['id']} was created, "
                                        f"but saving it to your account (Supabase) failed: {exc}\n\n"
                                        "This usually means SUPABASE_SERVICE_ROLE_KEY is missing, wrong, "
                                        "or swapped with the anon key in Streamlit Secrets."
                                    )

                if st.session_state.get("razorpay_checkout_url_free"):
                    st.link_button(
                        "💳 Continue to Razorpay Checkout",
                        st.session_state.razorpay_checkout_url_free,
                        use_container_width=True,
                    )
                    _free_sub_id = st.session_state.get("free_subscription_id", "")
                    if _free_sub_id:
                        st.caption(f"Subscription ID: {_free_sub_id}")
                        if st.button("🔄 Verify Payment", use_container_width=True, key=f"verify_free_{section_id}"):
                            try:
                                with st.spinner("Verifying your Razorpay subscription..."):
                                    active, status = verify_free_billing_subscription()
                                if active:
                                    st.success("✅ Payment verified. ₹299/mo billing is now active.")
                                    st.session_state.razorpay_checkout_url_free = ""
                                    st.rerun()
                                else:
                                    st.info(f"Payment is not active yet. Razorpay status: {status or 'unknown'}. Complete checkout and try again.")
                            except Exception as exc:
                                st.error(f"❌ Verification failed: {exc}")

            else:
                checkout_key = f"{name.upper()}_CHECKOUT_URL"
                checkout_url = secret(checkout_key)
                if checkout_url:
                    st.link_button(
                        button,
                        checkout_url,
                        use_container_width=True,
                    )
                else:
                    st.button(
                        button,
                        use_container_width=True,
                        disabled=True,
                        key=f"disabled_{section_id}_{name}",
                    )

    # Optional annual billing option — only shown if configured, so
    # nothing breaks for anyone who hasn't set this up yet.
    annual_url = secret("RAZORPAY_ANNUAL_CHECKOUT_URL")
    if annual_url:
        st.caption(
            "💡 Prefer to pay yearly? Save with annual billing on "
            "Professional."
        )
        st.link_button(
            "📅 See annual pricing",
            annual_url,
            use_container_width=False,
            key=f"annual_billing_{section_id}",
        )


# ============================================================
# AUTHENTICATION PAGE
# ============================================================

if not st.session_state.authenticated:

    # Topbar: logo + wordmark/tagline left, matching the reference design.
    # (The React reference shows "Need an account? Create account" here —
    # since Streamlit tabs don't support a clickable header-level tab
    # switch, this is a plain informational line instead.)
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), width=230)
    else:
        st.markdown(
            """<div class="gi-auth-topbar">
<div class="gi-brand-row">
<div class="gi-auth-icon-sm">✦</div>
<div>
<div class="gi-brand">Generative <span>Insight</span></div>
<div class="gi-tagline">AI Operations Copilot</div>
</div>
</div>
<div style="color:#737373; font-size:0.85rem;">Use the tabs below to sign in or create an account</div>
</div>""",
            unsafe_allow_html=True,
        )

    st.markdown(
        """<div class="gi-auth-hero" style="margin-top:18px;">
<div class="gi-auth-icon-lg">✦</div>
<div class="gi-auth-hero-title">AI-powered operational intelligence</div>
<div class="gi-auth-hero-sub">Turn operational data into management decisions. Create your account, upload Excel/CSV operational data, identify KPI risks, investigate team and employee performance, ask the AI Operations Copilot questions, and generate management-ready reports.</div>
</div>""",
        unsafe_allow_html=True,
    )

    if (
        not secret("SUPABASE_URL")
        or not secret("SUPABASE_ANON_KEY")
    ):
        st.error(
            "🔐 Authentication is not configured yet. "
            "Add SUPABASE_URL and SUPABASE_ANON_KEY in "
            "Streamlit → App Settings → Secrets."
        )
        st.stop()

    signup_tab, login_tab, pricing_tab = st.tabs(
        [
            "🆕 Create Account",
            "🔐 Sign In",
            "💳 Plans",
        ]
    )

    # --------------------------------------------------------
    # SIGN UP
    # --------------------------------------------------------

    with signup_tab:

        st.markdown(
            """<div class="gi-auth-card-title">Create your account</div>
<div class="gi-auth-card-desc">Start with the Free plan. You can upgrade later.</div>""",
            unsafe_allow_html=True,
        )

        with st.form(
            "signup_form",
            clear_on_submit=False,
        ):

            st.markdown('<div class="gi-field-label-row">👤 Full Name</div>', unsafe_allow_html=True)
            signup_name = st.text_input(
                "Full Name",
                placeholder="e.g. Sunil Sethy",
                label_visibility="collapsed",
            )

            st.markdown('<div class="gi-field-label-row">🏢 Company / Organization</div>', unsafe_allow_html=True)
            signup_company = st.text_input(
                "Company / Organization",
                placeholder="e.g. ABC Technologies",
                label_visibility="collapsed",
            )

            st.markdown('<div class="gi-field-label-row">✉️ Work Email</div>', unsafe_allow_html=True)
            signup_email = st.text_input(
                "Work Email",
                placeholder="name@company.com",
                label_visibility="collapsed",
            )

            st.markdown('<div class="gi-field-label-row">🔒 Password</div>', unsafe_allow_html=True)
            signup_password = st.text_input(
                "Password",
                type="password",
                placeholder="Use a strong password",
                help=(
                    "Use a strong password. Supabase enforces "
                    "the configured password policy."
                ),
                label_visibility="collapsed",
            )

            st.markdown('<div class="gi-field-label-row">🔒 Confirm Password</div>', unsafe_allow_html=True)
            signup_confirm = st.text_input(
                "Confirm Password",
                type="password",
                placeholder="Re-enter your password",
                label_visibility="collapsed",
            )

            signup_submitted = st.form_submit_button(
                "🚀 Create Free Account   →",
                type="primary",
                use_container_width=True,
            )

        st.markdown(
            """<div class="gi-auth-footer-note">
By creating an account, you agree to use the platform responsibly and validate AI recommendations before taking material business action.
</div>""",
            unsafe_allow_html=True,
        )

        if signup_submitted:

            if not signup_name.strip():
                st.warning(
                    "Please enter your full name."
                )

            elif not signup_company.strip():
                st.warning(
                    "Please enter your company or organization."
                )

            elif (
                not signup_email.strip()
                or "@" not in signup_email
            ):
                st.warning(
                    "Please enter a valid email address."
                )

            elif len(signup_password) < 6:
                st.warning(
                    "Please use a password with at least 6 characters."
                )

            elif signup_password != signup_confirm:
                st.warning(
                    "Passwords do not match."
                )

            else:

                try:

                    with st.spinner(
                        "Creating your account..."
                    ):

                        signup_response = sign_up_user(
                            signup_name,
                            signup_company,
                            signup_email,
                            signup_password,
                        )

                    signup_user = getattr(
                        signup_response,
                        "user",
                        None,
                    )

                    signup_session = getattr(
                        signup_response,
                        "session",
                        None,
                    )

                    if (
                        signup_user is not None
                        and signup_session is not None
                    ):

                        set_authenticated_user(
                            signup_response
                        )

                        st.success(
                            "✅ Account created successfully."
                        )

                        st.rerun()

                    elif signup_user is not None:

                        st.success(
                            "✅ Account created. Please check your "
                            "email and click the verification link "
                            "before signing in."
                        )

                    else:

                        st.info(
                            "If the email is valid, check your inbox "
                            "for the verification email."
                        )

                except Exception as e:

                    st.error(
                        "❌ Could not create account: "
                        + friendly_auth_error(e)
                    )

    # --------------------------------------------------------
    # LOGIN
    # --------------------------------------------------------

    with login_tab:

        st.markdown("### Welcome back")

        with st.form("login_form"):

            login_email = st.text_input(
                "Email",
                placeholder="name@company.com",
            )

            login_password = st.text_input(
                "Password",
                type="password",
            )

            login_submitted = st.form_submit_button(
                "🔐 Sign In",
                type="primary",
                use_container_width=True,
            )

        if login_submitted:

            if (
                not login_email.strip()
                or not login_password
            ):

                st.warning(
                    "Please enter your email and password."
                )

            else:

                try:

                    with st.spinner(
                        "Signing you in..."
                    ):

                        login_response = sign_in_user(
                            login_email,
                            login_password,
                        )

                    set_authenticated_user(
                        login_response
                    )

                    st.success(
                        "✅ Signed in successfully."
                    )

                    st.rerun()

                except Exception as e:

                    st.error(
                        "❌ Sign in failed: "
                        + friendly_auth_error(e)
                    )

        st.info(
            "If email confirmation is enabled in Supabase, "
            "verify your email before signing in."
        )

    # --------------------------------------------------------
    # PRICING ON LOGIN PAGE
    # --------------------------------------------------------

    with pricing_tab:
        show_pricing("login")

    st.stop()


# ============================================================
# SIDEBAR
# ============================================================

plan_config = get_plan_config(
    st.session_state.user_plan
)

with st.sidebar:

    if LOGO_PATH.exists():

        st.image(
            str(LOGO_PATH),
            use_container_width=True,
        )

    else:

        st.markdown(
            """<div class="gi-brand-row">
<div class="gi-logo-mark">✦</div>
<div class="gi-brand">Generative <span>Insight</span></div>
</div>""",
            unsafe_allow_html=True,
        )

    st.caption("AI Operations Copilot")

    st.markdown(
        f"""<a class="website-link" href="{WEBSITE_URL}" target="_blank" rel="noopener noreferrer">🌐 Visit Generative Insight</a>""",
        unsafe_allow_html=True,
    )

    st.divider()

    st.markdown(
        f'<div class="gi-pill green">Plan: {st.session_state.user_plan}</div>',
        unsafe_allow_html=True,
    )

    if st.session_state.user_plan == "Free":
        if free_billing_is_active():
            st.markdown(
                '<div class="gi-pill green">✅ ₹299/mo billing active — trial limits don\'t apply.</div>',
                unsafe_allow_html=True,
            )
        else:
            remaining = trial_days_remaining()
            if remaining is not None:
                if remaining <= 0:
                    st.markdown(
                        f'<div class="gi-pill red">Your {FREE_TRIAL_DAYS}-day free trial has ended. '
                        'Continue for ₹299/mo to keep using AI Operations Manager.</div>',
                        unsafe_allow_html=True,
                    )
                elif remaining <= 3:
                    st.markdown(
                        f'<div class="gi-pill amber">⏳ {remaining} day(s) left in your free trial '
                        '(then ₹299/mo).</div>',
                        unsafe_allow_html=True,
                    )
                elif remaining <= 5:
                    # Soft nudge before the hard wall hits — people convert
                    # better when they choose to upgrade early than when
                    # they're forced to at day 0.
                    st.markdown(
                        f'<div class="gi-pill amber">🙂 {remaining} days left in your trial. '
                        'Lock in ₹299/mo now to avoid any interruption.</div>',
                        unsafe_allow_html=True,
                    )
                    if st.button(
                        "Continue for ₹299/mo",
                        use_container_width=True,
                        key="sidebar_early_continue",
                    ):
                        st.session_state.show_plans = True
                        st.rerun()
                else:
                    st.markdown(
                        f'<div class="gi-pill neutral">{remaining} days left in your free trial (then ₹299/mo).</div>',
                        unsafe_allow_html=True,
                    )

    render_user_badge(st.session_state.get("user_name", ""), st.session_state.user_email)

    st.divider()

    st.header("🏭 Industry")

    _current_industry = st.session_state.get("industry", "BPO")
    _selected_label = st.selectbox(
        "Analyze data for",
        options=list(INDUSTRY_LABELS.values()),
        index=list(INDUSTRY_LABELS.keys()).index(_current_industry)
        if _current_industry in INDUSTRY_LABELS else 0,
        key="industry_selector",
    )
    _selected_industry = next(
        k for k, v in INDUSTRY_LABELS.items() if v == _selected_label
    )
    if _selected_industry != _current_industry:
        update_user_industry(_selected_industry)
        clear_analysis()
        st.session_state.manufacturing_file_name = ""
        st.session_state.manufacturing_result = None
        st.session_state.manufacturing_report_bytes = None
        st.rerun()

    st.divider()

    if st.session_state.industry == "BPO":

        st.header("⚙️ KPI Controls")

        productivity_target = st.number_input(
            "Productivity target %",
            min_value=1,
            max_value=200,
            value=90,
        )

        quality_target = st.number_input(
            "Quality target %",
            min_value=1,
            max_value=100,
            value=95,
        )

        sla_target = st.number_input(
            "SLA target %",
            min_value=1,
            max_value=100,
            value=97,
        )

        aht_target = st.number_input(
            "AHT target",
            min_value=1,
            max_value=1000,
            value=50,
        )

        st.divider()

    if st.button(
        "💳 View Plans",
        use_container_width=True,
        key="sidebar_view_plans",
    ):
        st.session_state.show_plans = True

    if st.button(
        "🔄 Reset Analysis",
        use_container_width=True,
        key="sidebar_reset_analysis",
    ):
        clear_analysis()
        st.rerun()

    if st.button(
        "🚪 Sign Out",
        use_container_width=True,
        key="sidebar_sign_out",
    ):
        clear_authentication()
        st.rerun()


if st.session_state.get("show_plans"):

    st.divider()

    show_pricing("sidebar")

    st.divider()


# ============================================================
# FREE TRIAL GATE
# 15 days of Free access, then the dashboard is locked until
# the user upgrades to a paid plan.
# ============================================================

if st.session_state.user_plan == "Free" and not free_billing_is_active():

    _trial_remaining = trial_days_remaining()

    if _trial_remaining is not None and _trial_remaining <= 0:

        show_brand_header(compact=True)

        st.markdown(
            '<div class="main-title">Your Free trial has ended</div>',
            unsafe_allow_html=True,
        )

        st.warning(
            f"Your {FREE_TRIAL_DAYS}-day free trial ended "
            f"{abs(_trial_remaining)} day(s) ago. Continue on Free for "
            "₹299/mo, or upgrade to Professional or Business, to keep "
            "analyzing operational data."
        )

        show_pricing("trial_expired")

        st.stop()

    elif _trial_remaining is not None and _trial_remaining <= 5:

        # Soft nudge before the hard wall — shown once trial is running
        # low but access is still allowed. Distinct from the sidebar
        # caption: this sits at the top of the main dashboard where it
        # can't be missed, with a direct link to upgrade now.
        nudge_col1, nudge_col2 = st.columns([4, 1])
        with nudge_col1:
            st.warning(
                f"⏳ **{_trial_remaining} day(s) left** in your free trial. "
                "Continue on Free for ₹299/mo, or upgrade to Professional, "
                "to avoid losing access to your dashboard."
            )
        with nudge_col2:
            if st.button(
                "View Plans",
                use_container_width=True,
                key="trial_nudge_view_plans",
            ):
                st.session_state.show_plans = True
                st.rerun()


# ============================================================
# HEADER
# ============================================================

show_brand_header(compact=True)

st.markdown(
    '<div class="main-title">AI Operations Manager</div>',
    unsafe_allow_html=True,
)

st.markdown(
    """<div class="brand-subtitle" style="color:#52637A !important; -webkit-text-fill-color:#52637A !important;">Executive operational intelligence → risk detection → AI decisions → action plans → management reports</div>""",
    unsafe_allow_html=True,
)

# ============================================================
# INDUSTRY DISPATCH
# Manufacturing gets its own complete, self-contained flow. Anything
# not yet in BUILT_INDUSTRIES gets the coming-soon screen. Both stop
# here. Everything below this point (Report Setup, file upload, KPI
# analysis, tabs) is BPO-only and assumes productivity_target etc.
# exist, which they only do when industry == "BPO".
# ============================================================

if st.session_state.industry == "Manufacturing":
    render_manufacturing_flow(plan_config)
    st.stop()

elif st.session_state.industry not in BUILT_INDUSTRIES:
    render_coming_soon_flow(st.session_state.industry)
    st.stop()


# ============================================================
# CUSTOMER INFORMATION
# ============================================================

st.subheader("🏢 Report Setup")

col1, col2, col3 = st.columns(3)

with col1:

    company_name = st.text_input(
        "Company Name",
        value=st.session_state.get(
            "company_name",
            "",
        ),
        placeholder="e.g. ABC Technologies",
    )

with col2:

    manager_email = st.text_input(
        "Manager Email",
        value=st.session_state.get(
            "user_email",
            "",
        ),
        placeholder="manager@company.com",
    )

with col3:

    report_name = st.text_input(
        "Report Name",
        value="Daily Operations Report",
    )


# ============================================================
# FILE UPLOAD
# ============================================================

uploaded = st.file_uploader(
    (
        "📁 Upload Excel or CSV operational data — "
        f"max {plan_config['max_mb']} MB"
    ),
    type=["xlsx", "xls", "csv"],
)

if not uploaded:

    st.info(
        "Upload operational data to activate the "
        "executive dashboard."
    )

    st.markdown("### Required columns")

    st.code(
        "Date, Employee_ID, Employee_Name, Team, Target, "
        "Production, AHT_Actual, AHT_Target, Quality_%, "
        "SLA_%, Attendance, Error_Count, Error_Category"
    )

    st.download_button(
        "⬇️ Download Data Template (.xlsx)",
        data=build_data_template_bytes(),
        file_name="AI_Operations_Manager_Data_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="download_data_template",
    )

    st.caption(
        "Includes example rows and a column-by-column guide — "
        "delete the sample rows and paste in your own data."
    )

    st.markdown("### What you get")

    a, b, c, d = st.columns(4)

    a.metric("Risk Detection", "✓")
    b.metric("Employee Risk", "✓")
    c.metric("AI Copilot", "✓")
    d.metric("Management Report", "✓")

    st.stop()


# ============================================================
# FILE SIZE
# ============================================================

file_mb = uploaded.size / (1024 * 1024)

if file_mb > plan_config["max_mb"]:

    st.error(
        f"File is {file_mb:.2f} MB. "
        f"Your {st.session_state.user_plan} plan supports "
        f"files up to {plan_config['max_mb']} MB."
    )

    st.stop()

elif (
    st.session_state.user_plan != "Professional"
    and file_mb > plan_config["max_mb"] * 0.8
):

    # Contextual upsell — shown right where the constraint actually
    # bites, not on a separate pricing page.
    st.warning(
        f"This file is {file_mb:.2f} MB, close to your "
        f"{plan_config['max_mb']} MB limit. Professional supports files "
        "up to 25 MB plus PDF + email reports and n8n automation."
    )
    if st.button("🚀 See Professional plan", key="upsell_filesize"):
        st.session_state.show_plans = True
        st.rerun()


# ============================================================
# RESET WHEN NEW FILE
# ============================================================

if st.session_state.file_name != uploaded.name:

    st.session_state.file_name = uploaded.name
    st.session_state.n8n_sent = False
    st.session_state.n8n_result = None
    st.session_state.copilot_answer = None
    st.session_state.last_question = ""
    st.session_state.analysis_result = None
    st.session_state.analysis_df = None
    st.session_state.report_pdf = None


# ============================================================
# N8N SETTINGS
# ============================================================

n8n_url = secret(
    "N8N_WEBHOOK_URL"
)

copilot_url = secret(
    "N8N_COPILOT_WEBHOOK_URL"
)

if n8n_url and "/webhook-test/" in n8n_url:

    st.warning(
        "⚠️ n8n is configured with a TEST webhook. "
        "For production use, activate the workflow and use "
        "/webhook/operations-upload in Streamlit Secrets."
    )

if (
    copilot_url
    and "/webhook-test/" in copilot_url
):

    st.warning(
        "⚠️ Management Copilot is using an n8n TEST webhook. "
        "Use the production /webhook/management-copilot URL "
        "after activating the workflow."
    )


# ============================================================
# READ FILE
# ============================================================

_analysis_start_time = time.time()

with st.spinner("🔄 Fetching and reading your uploaded data..."):

    try:

        uploaded.seek(0)

        if uploaded.name.lower().endswith(".csv"):

            df = pd.read_csv(uploaded)

        else:

            xls = pd.ExcelFile(uploaded)

            sheet = (
                "Operational_Data"
                if "Operational_Data" in xls.sheet_names
                else xls.sheet_names[0]
            )

            df = pd.read_excel(
                uploaded,
                sheet_name=sheet,
            )

    except Exception as e:

        st.error(
            f"❌ Could not read the uploaded file: {e}"
        )

        st.stop()


# ============================================================
# VALIDATE DATA
# ============================================================

required_columns = [
    "Employee_ID",
    "Employee_Name",
    "Team",
    "Target",
    "Production",
    "AHT_Actual",
    "Quality_%",
    "SLA_%",
    # These two matter: engine.py's analyze_data() requires them too
    # (see its REQUIRED list). Without checking for them here, a file
    # could pass this gate cleanly and then crash a moment later
    # inside analyze_data() with a confusing "Missing columns" error.
    "Attendance",
    "Error_Count",
]

missing_columns = [
    col
    for col in required_columns
    if col not in df.columns
]

if missing_columns:

    st.error(
        "❌ Required columns are missing."
    )

    st.write(missing_columns)

    st.stop()


# ============================================================
# LOCAL ANALYSIS
# ============================================================

with st.spinner("🧠 Analyzing operational data against your KPI targets..."):

    try:

        result = analyze_data(
            df,
            productivity_target=productivity_target,
            quality_target=quality_target,
            sla_target=sla_target,
            aht_target=aht_target,
        )

    except Exception as e:

        st.error(
            f"❌ Analysis failed: {e}"
        )

        st.stop()

_analysis_elapsed = time.time() - _analysis_start_time

st.session_state.analysis_result = result
st.session_state.analysis_df = df


# ============================================================
# KPI CALCULATIONS
# ============================================================

overall = result["overall"]

productivity = float(
    overall["productivity"]
)

quality = float(
    overall["quality"]
)

sla = float(
    overall["sla"]
)

aht = float(
    overall["aht"]
)

productivity_gap = (
    productivity - productivity_target
)

quality_gap = (
    quality - quality_target
)

sla_gap = (
    sla - sla_target
)

aht_gap = (
    aht - aht_target
)

breaches = sum(
    [
        productivity < productivity_target,
        quality < quality_target,
        sla < sla_target,
        aht > aht_target,
    ]
)

if breaches == 0:

    risk_level = "🟢 LOW RISK"

elif breaches == 1:

    risk_level = "🟡 MEDIUM RISK"

elif breaches == 2:

    risk_level = "🟠 HIGH RISK"

else:

    risk_level = "🔴 CRITICAL RISK"


actions_df = result.get(
    "actions",
    pd.DataFrame(),
)

action_count = (
    len(actions_df)
    if isinstance(actions_df, pd.DataFrame)
    else 0
)

high_priority_count = 0

if (
    isinstance(actions_df, pd.DataFrame)
    and not actions_df.empty
):

    for col in [
        "Priority",
        "priority",
        "Priority_Level",
        "priority_level",
    ]:

        if col in actions_df.columns:

            high_priority_count = len(
                actions_df[
                    actions_df[col]
                    .astype(str)
                    .str.lower()
                    .isin(
                        [
                            "high",
                            "critical",
                        ]
                    )
                ]
            )

            break


# ============================================================
# EXECUTIVE HERO
# ============================================================

st.markdown(
    f"""<div class="hero">
<h3>Executive Health: {risk_level}</h3>
<p class="small-muted">{company_name or "Your organization"} · {report_name}</p>
<p class="small-muted">⏱️ Data fetched and analyzed in {_analysis_elapsed:.2f}s · {len(df)} row(s) processed</p>
</div>""",
    unsafe_allow_html=True,
)


# ============================================================
# TOP EXECUTIVE METRICS
# ============================================================

r1, r2, r3, r4 = st.columns(4)

with r1:
    render_metric_card("🚦", "Operational Risk", risk_level)

with r2:
    render_metric_card(
        "🎯", "KPI Breaches", breaches,
        delta=f"{4 - breaches} on target",
        delta_tone="good" if breaches == 0 else ("bad" if breaches >= 2 else "neutral"),
    )

with r3:
    render_metric_card("✅", "Action Items", action_count)

with r4:
    render_metric_card(
        "🔺", "High/Critical Actions", high_priority_count,
        delta_tone="bad" if high_priority_count > 0 else "good",
    )


# ============================================================
# KPI PERFORMANCE
# ============================================================

st.subheader(
    "📊 KPI Performance vs Target"
)

k1, k2, k3, k4 = st.columns(4)

with k1:
    render_metric_card(
        "📈", "Productivity", f"{productivity:.2f}%",
        delta=f"{productivity_gap:+.2f}% vs target",
        delta_tone="good" if productivity_gap >= 0 else "bad",
    )

with k2:
    render_metric_card(
        "✔️", "Quality", f"{quality:.2f}%",
        delta=f"{quality_gap:+.2f}% vs target",
        delta_tone="good" if quality_gap >= 0 else "bad",
    )

with k3:
    render_metric_card(
        "⏱️", "SLA", f"{sla:.2f}%",
        delta=f"{sla_gap:+.2f}% vs target",
        delta_tone="good" if sla_gap >= 0 else "bad",
    )

with k4:
    render_metric_card(
        "⌛", "Average AHT", f"{aht:.2f}",
        delta=f"{aht_gap:+.2f} vs target",
        delta_tone="bad" if aht_gap > 0 else "good",
    )


# ============================================================
# EXECUTIVE SUMMARY
# ============================================================

summary_points = []

summary_points.append(
    f"{'🔴' if productivity < productivity_target else '🟢'} "
    f"Productivity: {productivity:.1f}% vs "
    f"{productivity_target}% target."
)

summary_points.append(
    f"{'🔴' if quality < quality_target else '🟢'} "
    f"Quality: {quality:.1f}% vs "
    f"{quality_target}% target."
)

summary_points.append(
    f"{'🔴' if sla < sla_target else '🟢'} "
    f"SLA: {sla:.1f}% vs "
    f"{sla_target}% target."
)

summary_points.append(
    f"{'🟠' if aht > aht_target else '🟢'} "
    f"AHT: {aht:.1f} vs "
    f"{aht_target} target."
)

if breaches >= 3:

    recommendation = (
        "Immediate management attention is recommended. "
        "Multiple KPI thresholds are breached. Prioritize "
        "root-cause analysis, targeted corrective actions, "
        "and close monitoring."
    )

elif breaches >= 1:

    recommendation = (
        "Management should review the affected KPIs, validate "
        "contributing factors, and initiate targeted corrective actions."
    )

else:

    recommendation = (
        "Operations are within defined KPI thresholds. Continue "
        "monitoring performance and maintain current processes."
    )


st.subheader("🧠 Executive Summary")

for point in summary_points:
    st.write(point)

st.info(
    f"💡 **Management Recommendation:** {recommendation}"
)


# ============================================================
# N8N OPERATIONAL AUTOMATION
# ============================================================

if n8n_url and not st.session_state.n8n_sent:

    if (
        not company_name.strip()
        or not manager_email.strip()
    ):

        st.warning(
            "Enter Company Name and Manager Email to run "
            "the configured n8n operational automation."
        )

    else:

        try:

            uploaded.seek(0)

            files = {
                "file": (
                    uploaded.name,
                    uploaded.getvalue(),
                    uploaded.type
                    or "application/octet-stream",
                )
            }

            data = {
                "company_name": company_name.strip(),
                "manager_email": manager_email.strip(),
                "report_name": report_name.strip(),
                "user_id": st.session_state.user_id,
                "user_email": st.session_state.user_email,
            }

            with st.spinner(
                "🤖 Running operational automation..."
            ):

                response = requests.post(
                    n8n_url,
                    files=files,
                    data=data,
                    timeout=120,
                )

            if response.status_code < 300:

                st.session_state.n8n_result = (
                    normalize_n8n_response(response)
                )

                st.session_state.n8n_sent = True

                st.success(
                    "✅ Operational automation completed."
                )

            else:

                st.error(
                    f"❌ n8n workflow failed: "
                    f"HTTP {response.status_code}"
                )

                st.code(
                    response.text,
                    language="text",
                )

        except requests.exceptions.Timeout:

            st.warning(
                "⏱️ n8n timed out. The workflow may still be running."
            )

        except requests.exceptions.RequestException as e:

            st.error(
                f"❌ Could not connect to n8n: {e}"
            )


# ============================================================
# MAIN TABS
# ============================================================

tabs = st.tabs(
    [
        "📊 Executive Dashboard",
        "🚨 AI Insights",
        "👥 Employee Risk",
        "✅ Action Center",
        "🤖 Management Copilot",
        "📄 Reports",
        "💳 Billing",
    ]
)


# ============================================================
# TAB 1 — EXECUTIVE DASHBOARD
# ============================================================

with tabs[0]:

    left, right = st.columns([1.4, 1])

    with left:

        st.subheader("Team Performance")

        team_df = result.get(
            "team",
            pd.DataFrame(),
        )

        if (
            isinstance(team_df, pd.DataFrame)
            and not team_df.empty
        ):

            st.dataframe(
                team_df,
                use_container_width=True,
                hide_index=True,
            )

            if (
                "Team" in team_df.columns
                and "Productivity_%" in team_df.columns
            ):

                st.subheader(
                    "Productivity by Team"
                )

                st.bar_chart(
                    team_df.set_index("Team")[
                        "Productivity_%"
                    ]
                )

        else:

            st.info(
                "No team-level data available."
            )

    with right:

        st.subheader(
            "Management Snapshot"
        )

        employees = result.get(
            "employees",
            pd.DataFrame(),
        )

        if (
            isinstance(employees, pd.DataFrame)
            and not employees.empty
        ):

            st.metric(
                "Employees analyzed",
                len(employees),
            )

            if "Risk_Score" in employees.columns:

                st.metric(
                    "Highest employee risk score",
                    f"{employees['Risk_Score'].max():.2f}",
                )

        st.write(
            "**Current KPI position**"
        )

        for item in summary_points:
            st.write(item)


# ============================================================
# TAB 2 — AI INSIGHTS
# ============================================================

with tabs[1]:

    st.subheader(
        "🚨 Automated Findings"
    )

    findings_df = result.get(
        "findings",
        pd.DataFrame(),
    )

    if (
        isinstance(findings_df, pd.DataFrame)
        and not findings_df.empty
    ):

        st.dataframe(
            findings_df,
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.success(
            "✅ No threshold breaches detected."
        )

    st.info(
        "Root causes are evidence-based hypotheses. "
        "The available data may not prove causality."
    )


# ============================================================
# TAB 3 — EMPLOYEE RISK
# ============================================================

with tabs[2]:

    st.subheader(
        "👥 Employee Risk"
    )

    employee_data = result.get(
        "employees",
        pd.DataFrame(),
    )

    if (
        isinstance(employee_data, pd.DataFrame)
        and not employee_data.empty
    ):

        sort_cols = [
            c
            for c in [
                "Risk_Score",
                "Avg_Productivity",
            ]
            if c in employee_data.columns
        ]

        if sort_cols:

            employee_data = employee_data.sort_values(
                sort_cols,
                ascending=[
                    False
                ] * len(sort_cols),
            )

        st.dataframe(
            employee_data,
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.info(
            "No employee-level risk data available."
        )


# ============================================================
# TAB 4 — ACTION CENTER
# ============================================================

with tabs[3]:

    st.subheader(
        "✅ Recommended Actions"
    )

    if (
        isinstance(actions_df, pd.DataFrame)
        and not actions_df.empty
    ):

        st.dataframe(
            actions_df,
            use_container_width=True,
            hide_index=True,
        )

    else:

        st.success(
            "No action items generated."
        )


# ============================================================
# TAB 5 — MANAGEMENT COPILOT
# ============================================================

with tabs[4]:

    st.subheader(
        "🤖 Management Copilot"
    )

    st.caption(
        "Ask questions about the uploaded operational data. "
        "The Copilot is instructed to use only the supplied "
        "operational context."
    )

    question = st.text_input(
        "Ask your operational question",
        placeholder=(
            "Which team has the quality drop and "
            "what action should be taken?"
        ),
        key="copilot_question",
    )

    ask_copilot = st.button(
        "🚀 Ask Management Copilot",
        type="primary",
        use_container_width=True,
        key="ask_management_copilot",
    )

    if ask_copilot:

        if not question.strip():

            st.warning(
                "⚠️ Please enter a question first."
            )

        elif not plan_config["copilot"]:

            st.error(
                "Copilot is not available on this plan."
            )

        elif not copilot_url:

            st.error(
                "❌ N8N_COPILOT_WEBHOOK_URL is not "
                "configured in Streamlit Secrets."
            )

        else:

            context = build_copilot_context(
                company_name,
                report_name,
                result,
                productivity_target,
                quality_target,
                sla_target,
                aht_target,
                risk_level,
                summary_points,
            )

            payload = {
                "question": question.strip(),
                "company_name": company_name.strip(),
                "report_name": report_name.strip(),
                "context": context,
                "user_id": st.session_state.user_id,
                "user_email": st.session_state.user_email,
            }

            try:

                with st.spinner(
                    "🤖 Management Copilot is analyzing..."
                ):

                    copilot_response = requests.post(
                        copilot_url,
                        json=payload,
                        headers={
                            "Content-Type": "application/json"
                        },
                        timeout=120,
                    )

                if copilot_response.status_code < 300:

                    raw = normalize_n8n_response(
                        copilot_response
                    )

                    answer_data = parse_ai_answer(
                        raw
                    )

                    st.session_state.copilot_answer = (
                        answer_data
                    )

                    st.session_state.last_question = (
                        question.strip()
                    )

                else:

                    st.session_state.copilot_answer = None

                    st.error(
                        "❌ Copilot workflow failed: "
                        f"HTTP {copilot_response.status_code}"
                    )

                    st.code(
                        copilot_response.text,
                        language="text",
                    )

            except requests.exceptions.Timeout:

                st.session_state.copilot_answer = None

                st.error(
                    "⏱️ Management Copilot timed out. "
                    "Please try again."
                )

            except requests.exceptions.ConnectionError:

                st.session_state.copilot_answer = None

                st.error(
                    "🔌 Could not connect to the n8n "
                    "Copilot webhook."
                )

            except requests.exceptions.RequestException as e:

                st.session_state.copilot_answer = None

                st.error(
                    f"❌ Copilot request failed: {e}"
                )

            except Exception as e:

                st.session_state.copilot_answer = None

                st.error(
                    f"❌ Unexpected Copilot error: {e}"
                )

    if st.session_state.copilot_answer:

        st.divider()

        st.markdown(
            "### 🧠 Copilot Analysis"
        )

        st.caption(
            "Question: "
            + st.session_state.last_question
        )

        answer = st.session_state.copilot_answer

        if isinstance(answer, dict):

            what = answer.get(
                "what_is_happening"
            )

            factors = answer.get(
                "contributing_factors",
                [],
            )

            rec_actions = answer.get(
                "recommended_actions",
                [],
            )

            priority = answer.get(
                "priority",
                "",
            )

            owner = answer.get(
                "owner",
                "",
            )

            timeline = answer.get(
                "timeline",
                "",
            )

            sufficiency = answer.get(
                "data_sufficiency"
            )

            if what:

                st.markdown(
                    "#### 🔎 What is happening"
                )

                st.info(what)

            if factors:

                st.markdown(
                    "#### 🔍 Contributing Factors"
                )

                for factor in factors:
                    st.write(
                        f"• {factor}"
                    )

            if rec_actions:

                st.markdown(
                    "#### ✅ Recommended Actions"
                )

                for i, action in enumerate(
                    rec_actions,
                    1,
                ):

                    st.markdown(
                        f"**{i}.** {action}"
                    )

            st.markdown(
                "#### 📌 Management Decision"
            )

            d1, d2, d3 = st.columns(3)

            with d1:
                st.metric(
                    "Priority",
                    priority or "N/A",
                )

            with d2:
                st.metric(
                    "Owner",
                    owner or "N/A",
                )

            with d3:
                st.metric(
                    "Timeline",
                    timeline or "N/A",
                )

            if sufficiency:

                st.markdown(
                    "#### 📊 Data Sufficiency"
                )

                st.warning(
                    sufficiency
                )

        elif isinstance(answer, str):

            st.markdown(answer)

        else:

            st.code(
                str(answer),
                language="text",
            )


# ============================================================
# TAB 6 — REPORTS
# ============================================================

with tabs[5]:

    st.subheader(
        "📄 Management Reports"
    )

    if not plan_config["pdf"]:

        st.warning(
            "PDF reporting is not available on your current plan."
        )

    else:

        report_result = dict(result)

        report_result["_targets"] = {
            "productivity": productivity_target,
            "quality": quality_target,
            "sla": sla_target,
            "aht": aht_target,
        }

        if st.button(
            "📄 Generate Executive PDF",
            type="primary",
            use_container_width=True,
            key="generate_executive_pdf",
        ):

            try:

                with st.spinner(
                    "Generating management report..."
                ):

                    pdf = create_pdf_report(
                        company_name
                        or "Organization",
                        report_name
                        or "Operations Report",
                        report_result,
                        risk_level,
                        summary_points,
                        recommendation,
                    )

                st.session_state.report_pdf = pdf

                st.session_state.report_generated_at = (
                    datetime.now()
                )

            except Exception as e:

                st.error(
                    f"❌ Could not generate PDF: {e}"
                )

        if st.session_state.report_pdf:

            filename = (
                f"{company_name or 'operations'}_report_"
                f"{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            )

            st.download_button(
                "⬇️ Download Executive PDF",
                data=st.session_state.report_pdf,
                file_name=filename,
                mime="application/pdf",
                use_container_width=True,
                key="download_executive_pdf",
            )

            if st.session_state.report_generated_at:

                st.caption(
                    "Generated "
                    + st.session_state.report_generated_at.strftime(
                        "%d %b %Y, %H:%M"
                    )
                )

        st.divider()

        st.subheader(
            "📧 Email Report"
        )

        if not plan_config["email"]:

            st.info(
                "Email delivery is available on "
                "Professional and Business plans."
            )

        else:

            recipient = st.text_input(
                "Recipient email",
                value=manager_email,
                key="report_recipient",
            )

            if st.button(
                "📨 Email PDF Report",
                use_container_width=True,
                key="email_pdf_report",
            ):

                if not recipient.strip():

                    st.warning(
                        "Enter a recipient email address."
                    )

                elif not st.session_state.report_pdf:

                    st.warning(
                        "Generate the PDF first."
                    )

                else:

                    try:

                        send_email_report(
                            recipient.strip(),
                            f"{company_name} - {report_name}",
                            (
                                "Please find attached the management "
                                f"report for {company_name or 'your organization'}.\n\n"
                                "Generated by Generative Insight AI "
                                "Operations Copilot."
                            ),
                            st.session_state.report_pdf,
                            "operations_report.pdf",
                        )

                        st.success(
                            "✅ Report emailed successfully."
                        )

                    except Exception as e:

                        st.error(
                            f"❌ Email failed: {e}"
                        )

        st.divider()

        st.subheader(
            "📥 Data Exports"
        )

        e1, e2 = st.columns(2)

        with e1:

            team_df = result.get(
                "team",
                pd.DataFrame(),
            )

            if isinstance(
                team_df,
                pd.DataFrame,
            ):

                st.download_button(
                    "⬇️ Team Analysis CSV",
                    team_df.to_csv(
                        index=False
                    ).encode("utf-8"),
                    "team_analysis.csv",
                    "text/csv",
                    use_container_width=True,
                    key="download_team_analysis",
                )

        with e2:

            if isinstance(
                actions_df,
                pd.DataFrame,
            ):

                st.download_button(
                    "⬇️ Action Plan CSV",
                    actions_df.to_csv(
                        index=False
                    ).encode("utf-8"),
                    "action_plan.csv",
                    "text/csv",
                    use_container_width=True,
                    key="download_action_plan",
                )


# ============================================================
# TAB 7 — BILLING
# ============================================================

with tabs[6]:

    st.subheader(
        "💳 Subscription & Billing"
    )

    st.info(
        f"You are currently using the "
        f"**{st.session_state.user_plan}** plan."
    )

    show_pricing("billing")

    st.caption("Professional subscriptions are processed through Razorpay. Complete checkout and verify payment to activate access.")


# ============================================================
# AI PROMPT / DEBUG AREA
# ============================================================

with st.expander(
    "🧠 AI Analyst Context / Prompt",
    expanded=False,
):

    st.code(
        make_ai_prompt(result),
        language="text",
    )


# ============================================================
# FOOTER
# ============================================================

st.divider()

st.markdown(
    f"""<div class="gi-footer">
<strong>Generative Insight</strong> · AI Operations Copilot v{APP_VERSION}
<br>
Insights today. Intelligence tomorrow.
<br>
<a class="website-link" href="{WEBSITE_URL}" target="_blank" rel="noopener noreferrer">generativeinsight.in</a>
&nbsp;·&nbsp;
© {datetime.now().year}
</div>""",
    unsafe_allow_html=True,
)
